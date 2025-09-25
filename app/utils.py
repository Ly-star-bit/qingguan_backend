from copy import copy
from datetime import date, datetime, timedelta
import io
import os
from pathlib import Path
import random

import textwrap
import time
import traceback
import zipfile

from bson import ObjectId
import httpx
import jwt
import openpyxl
from openpyxl.styles import Alignment

from typing import List, Optional

import requests
from sqlmodel import  Session, select

from loguru import logger
from dotenv import load_dotenv

from morelink_api import MoreLinkClient
from rpa_tools.email_tools import send_email
from .schemas import ProductData
from .db import pool_engine
import jpype

from datetime import datetime, timedelta
import pandas as pd
# from pymongo import MongoClient
from fastapi import HTTPException
from openpyxl import Workbook as Openpyxl_Workbook
from app.db_mongo import get_db
from rpa_tools import find_playwright_node_path
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import  UploadFile
from fastapi.responses import StreamingResponse
from minio import Minio
from minio.error import S3Error
import PyPDF2
import re
import json
load_dotenv()


# 启动 JVM 并确保 JVM 启动在导入之前

# if not jpype.isJVMStarted():
#         jpype.startJVM()

# 在 JVM 启动后导入 Java 依赖的模块
# from asposecells.api import Workbook, License, PdfSaveOptions,TextAlignmentType, SaveFormat,SheetSet
# 密钥和算法
ACCESS_TOKEN_SECRET_KEY = os.getenv("ACCESS_TOKEN_SECRET_KEY")
ACCESS_TOKEN_ALGORITHM = os.getenv("ACCESS_TOKEN_ALGORITHM")
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now() + expires_delta
    else:
        expire = datetime.now() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, ACCESS_TOKEN_SECRET_KEY, algorithm=ACCESS_TOKEN_ALGORITHM)
    return encoded_jwt

def create_refresh_token(data: dict):
    expire = datetime.now() + timedelta(days=1)
    to_encode = data.copy()
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, ACCESS_TOKEN_SECRET_KEY, algorithm=ACCESS_TOKEN_ALGORITHM)
    return encoded_jwt
# def set_cell_value(sheet, row, column, value):
#         cell = sheet.getCells().get(row - 1, column - 1)  # Adjust for zero-based index
        
#         if isinstance(value, str) and value.startswith('='):
#             # 如果 value 是一个公式
#             cell.setFormula(value)
#         else:
#             # 否则，设置为普通值
#             cell.putValue(value)
        
#         style = cell.getStyle()
#         style.setTextWrapped(True)
#         style.setHorizontalAlignment(TextAlignmentType.CENTER)
#         style.setVerticalAlignment(TextAlignmentType.CENTER)
#         cell.setStyle(style)
# def generate_fencangdan_file(data):
#     apcelllic = License()
#     apcelllic.setLicense('JAVA-Aspose.Excel-24.7/license.xml')
#     fendan_path = r".\file\excel_template\分单模板 - 执行.xlsx"
#     cangdan_path = r".\file\excel_template\舱单模板 - 执行.xlsx"

#     cangdan_wb = Workbook(cangdan_path)
#     cangdan_sheet = cangdan_wb.getWorksheets().get("Sheet1")
#     #处理舱单
#     set_cell_value(cangdan_sheet,3,9,data['orderNumber'])
#     set_cell_value(cangdan_sheet,5,9,data['flight_no'])

#     cangdan_row = 11
#     for suborder in data['subOrders']:
#         set_cell_value(cangdan_sheet,cangdan_row,1,suborder['subOrderNumber'])
#         set_cell_value(cangdan_sheet,cangdan_row,3,suborder['boxCount'])
#         set_cell_value(cangdan_sheet,cangdan_row,4,suborder['grossWeight'])
#         set_cell_value(cangdan_sheet,cangdan_row,5,data['startland'])
#         set_cell_value(cangdan_sheet,cangdan_row,6,data['destination'])
#         set_cell_value(cangdan_sheet,cangdan_row,7,suborder['sender'])
#         set_cell_value(cangdan_sheet,cangdan_row,8,suborder['receiver'])

#         cangdan_sheet.autoFitRow(cangdan_row - 1)
#         cangdan_row += 1


#     output_dir = Path("file/fencangdan/cangdan")
#     output_dir.mkdir(parents=True, exist_ok=True)
#     output_path = str(output_dir / f"{time.time()}-{data['orderNumber']} .xlsx")

#     cangdan_wb.save(output_path, SaveFormat.XLSX)
#     print(f"Excel file generated: {output_path}")

#     # 生成 PDF 文件
#     pdf_path = excel2pdf(output_path, 'pdf')
#     return pdf_path
#     #处理分单
#     # fendan_wb =    openpyxl.load_workbook(fendan_path)
#     # fendan_sheet = fendan_wb.active
#     # for suborder in data['subOrders']:
#     #     set_cell_value(fendan_sheet,1,1,data['orderNumber'])
#     #     set_cell_value(fendan_sheet,1,35,suborder['subOrderNumber'])
#     #     set_cell_value(fendan_sheet,4,1,suborder['sender'])
#     #     set_cell_value(fendan_sheet,5,1,suborder['receiver'])
#     #     set_cell_value(fendan_sheet,15,1,data['startland'])
#     #     set_cell_value(fendan_sheet,18,1,data[''])
#     #     set_cell_value(fendan_sheet,20,1,data['destination'])
#     #     set_cell_value(fendan_sheet,20,9,data['flight_no'])
def generate_admin_shenhe_template(data, totalyugutax):
    start_time = time.time()
    template_path = "./file/excel_template/ADMIN-审核文件模板-0411.xlsx"
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # Get the style from the template row (row 14) to copy to new rows
    template_row = 14
    template_styles = {}
    for col in range(1, sheet.max_column + 1):
        template_cell = sheet.cell(row=template_row, column=col)
        template_styles[col] = {
            'font': copy(template_cell.font),
            'border': copy(template_cell.border),
            'fill': copy(template_cell.fill),
            'number_format': template_cell.number_format,
            'protection': copy(template_cell.protection),
            'alignment': copy(template_cell.alignment)
        }
    
    # Generate document info
    today_minus_5 = datetime.now() - timedelta(days=5)
    formatted_date = today_minus_5.strftime("%Y%m%d")
    random_number = random.randint(1000, 9999)
    result_1 = f"{formatted_date}{random_number}"
    
    # Fill header information
    sheet.cell(row=1, column=1).value = data[0]["shipper_name"]
    sheet.cell(row=2, column=1).value = data[0]["shipper_address"]
    sheet.cell(row=6, column=1).value = f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}"
    sheet.cell(row=6, column=14).value = result_1
    sheet.cell(row=7, column=14).value = result_1
    sheet.cell(row=8, column=14).value = datetime.now().strftime("%Y/%m/%d")
    sheet.cell(row=10, column=14).value = data[0]["MasterBillNo"]
    
    # Fill item data
    for index, item in enumerate(data):
        civ_row = 14 + index
        
        # Set values
        sheet.cell(row=civ_row, column=1).value = index + 1
        sheet.cell(row=civ_row, column=2).value = item["HS_CODE"]
        sheet.cell(row=civ_row, column=3).value = item["duty"]
        sheet.cell(row=civ_row, column=4).value = item["additional_duty"]
        sheet.cell(row=civ_row, column=5).value = item["DESCRIPTION"]
        sheet.cell(row=civ_row, column=6).value = item["ChineseName"]
        sheet.cell(row=civ_row, column=7).value = f"=I{civ_row}*J{civ_row}"
        sheet.cell(row=civ_row, column=8).value = item["danwei"]
        sheet.cell(row=civ_row, column=9).value = int(item["quanity"] / item["carton"])
        sheet.cell(row=civ_row, column=10).value = item["carton"]
        sheet.cell(row=civ_row, column=11).value = item["unit_price"]
        sheet.cell(row=civ_row, column=12).value = f"=K{civ_row}*G{civ_row}"
        sheet.cell(row=civ_row, column=13).value = item["texture"]
        sheet.cell(row=civ_row, column=14).value = item["note"]
        sheet.cell(row=civ_row, column=15).value = f"=round(P{civ_row}*0.8,2)"
        sheet.cell(row=civ_row, column=16).value = item['GrossWeight']
        sheet.cell(row=civ_row, column=17).value = item['Volume']

        if item.get("single_weight"):
             #颜色设置为红色
             sheet.cell(row=civ_row, column=16).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            #
        # 处理additional_duty为百分比或小数的情况
        def convert_duty_value(duty_str):
            if isinstance(duty_str, str) and duty_str.endswith('%'):
                try:
                    return float(duty_str.strip('%')) / 100
                except Exception:
                    return 0
            else:
                try:
                    return float(duty_str)
                except Exception:
                    return 0
                    
        additional_duty_value = convert_duty_value(item["additional_duty"])
        duty_value = convert_duty_value(item["duty"])
        sheet.cell(row=civ_row, column=18).value = f"=round(D{civ_row}*L{civ_row},2)"
        sheet.cell(row=civ_row, column=19).value = f"=round(C{civ_row}*L{civ_row},2)"
        # Insert new row below and copy styles
        sheet.insert_rows(civ_row + 1)
        
        # Apply styles to the new row
        for col in range(1, sheet.max_column + 1):
            new_cell = sheet.cell(row=civ_row + 1, column=col)
            if col in template_styles:
                style = template_styles[col]
                new_cell.font = style['font']
                new_cell.border = style['border']
                new_cell.fill = style['fill']
                new_cell.number_format = style['number_format']
                new_cell.protection = style['protection']
                new_cell.alignment = style['alignment']

    # 设置求和公式，从14行开始到当前行
    sheet.cell(row=civ_row+3, column=7).value = f"=SUM(G14:G{civ_row})"
    sheet.cell(row=civ_row+3, column=10).value = f"=SUM(J14:J{civ_row})"
    sheet.cell(row=civ_row+3, column=12).value = f"=SUM(L14:L{civ_row})"
    sheet.cell(row=civ_row+3, column=15).value = f"=SUM(O14:O{civ_row})"
    sheet.cell(row=civ_row+3, column=16).value = f"=SUM(P14:P{civ_row})"
    sheet.cell(row=civ_row+3, column=17).value = f"=SUM(Q14:Q{civ_row})"
    sheet.cell(row=civ_row+3, column=18).value = f"=SUM(R14:R{civ_row})"
    sheet.cell(row=civ_row+3, column=19).value = f"=SUM(S14:S{civ_row})"
    # logger.info(f"data: {data}")
    average_single_weight = sum([item.get("GrossWeight") for item in data if not item.get("single_weight")]) /  sum([item.get("carton") for item in data if not item.get("single_weight")])
    #添加单箱重量
    sheet.cell(row=civ_row+5, column=15).value = "单箱重量: "
    sheet.cell(row=civ_row+5, column=16).value = f"{round(average_single_weight,2)}"

    sheet.cell(row=civ_row+6, column=15).value = "预估总税金: "
    sheet.cell(row=civ_row+6, column=16).value = f"=round({data[0].get('estimated_tax_amount','')},2)"
    sheet.cell(row=civ_row+7, column=15).value = "货值比: "
    sheet.cell(row=civ_row+7, column=16).value = f"=round(L{civ_row+3}/P{civ_row+3},2)"

    sheet.cell(row=civ_row+8, column=15).value = "美国税率: "
    sheet.cell(row=civ_row+8, column=16).value = f"{data[0].get('rate','')}"


    sheet.cell(row=civ_row+9, column=15).value = "税金单价: "
    sheet.cell(row=civ_row+9, column=16).value = f"{data[0].get('estimated_tax_rate_cny_per_kg','')}"
    # 保存文件
    output_path = f"file/{time.time()}-{data[0]['MasterBillNo']} CI&PL-{totalyugutax}_admin_审核.xlsx"
    wb.save(output_path)
    end_time = time.time()
    print(f"shenzhen_customes_pdf_gennerate 审核模板 运行时间: {end_time - start_time:.2f} 秒")
    return output_path
def generate_admin_shenhe_canada_template(data, totalyugutax):
    template_path = "./file/excel_template/加拿大_admin_审核-模板-0606.xlsx"
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # Get the style from the template row (row 13) to copy to new rows
    template_row = 13
    template_styles = {}
    for col in range(1, sheet.max_column + 1):
        template_cell = sheet.cell(row=template_row, column=col)
        template_styles[col] = {
            'font': copy(template_cell.font),
            'border': copy(template_cell.border),
            'fill': copy(template_cell.fill),
            'number_format': template_cell.number_format,
            'protection': copy(template_cell.protection),
            'alignment': copy(template_cell.alignment)
        }
    
    # Generate document info
    today_minus_5 = datetime.now() - timedelta(days=5)
    formatted_date = today_minus_5.strftime("%Y%m%d")
    random_number = random.randint(1000, 9999)
    result_1 = f"{formatted_date}{random_number}"
    
    # Fill header information
    # sheet.cell(row=1, column=1).value = data[0]["shipper_name"]
    # sheet.cell(row=2, column=1).value = data[0]["shipper_address"]
    # sheet.cell(row=6, column=1).value = f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}"
    sheet.cell(row=5, column=14).value = result_1
    sheet.cell(row=6, column=14).value = result_1
    sheet.cell(row=7, column=14).value = datetime.now().strftime("%Y/%m/%d")
    sheet.cell(row=9, column=14).value = data[0]["MasterBillNo"]
    
    # Fill item data
    for index, item in enumerate(data):
        civ_row = 13 + index
        
        # Set values
        sheet.cell(row=civ_row, column=1).value = index + 1
        sheet.cell(row=civ_row, column=2).value = item["HS_CODE"]
        sheet.cell(row=civ_row, column=3).value = item["duty"]
        sheet.cell(row=civ_row, column=4).value = item["additional_duty"]
        sheet.cell(row=civ_row, column=5).value = item["DESCRIPTION"]
        sheet.cell(row=civ_row, column=6).value = item["ChineseName"]
        sheet.cell(row=civ_row, column=7).value = item["quanity"]
        sheet.cell(row=civ_row, column=8).value = item["danwei"]
        sheet.cell(row=civ_row, column=9).value = int(item["quanity"] / item["carton"])
        sheet.cell(row=civ_row, column=10).value = item["carton"]
        sheet.cell(row=civ_row, column=11).value = item["unit_price"]
        sheet.cell(row=civ_row, column=12).value = item["total_price"]
        sheet.cell(row=civ_row, column=13).value = item["texture"]
        sheet.cell(row=civ_row, column=14).value = item["note"]
        sheet.cell(row=civ_row, column=15).value = item["net_weight"]
        sheet.cell(row=civ_row, column=16).value = item["GrossWeight"]
        sheet.cell(row=civ_row, column=17).value = item["Volume"]

        if item.get("single_weight"):
             #颜色设置为红色
             sheet.cell(row=civ_row, column=16).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            #
        # 处理additional_duty为百分比或小数的情况
        def convert_duty_value(duty_str):
            if isinstance(duty_str, str) and duty_str.endswith('%'):
                try:
                    return float(duty_str.strip('%')) / 100
                except Exception:
                    return 0
            else:
                try:
                    return float(duty_str)
                except Exception:
                    return 0
                    
        additional_duty_value = convert_duty_value(item["additional_duty"])
        duty_value = convert_duty_value(item["duty"])
        sheet.cell(row=civ_row, column=18).value = round(float(item["total_price"])) * additional_duty_value
        sheet.cell(row=civ_row, column=19).value = round(float(item["total_price"])) * duty_value
        # Insert new row below and copy styles
        sheet.insert_rows(civ_row + 1)
        
        # Apply styles to the new row
        for col in range(1, sheet.max_column + 1):
            new_cell = sheet.cell(row=civ_row + 1, column=col)
            if col in template_styles:
                style = template_styles[col]
                new_cell.font = style['font']
                new_cell.border = style['border']
                new_cell.fill = style['fill']
                new_cell.number_format = style['number_format']
                new_cell.protection = style['protection']
                new_cell.alignment = style['alignment']

    # 设置求和公式，从14行开始到当前行
    # sheet.cell(row=civ_row+3, column=7).value = f"=SUM(G14:G{civ_row})"
    # sheet.cell(row=civ_row+3, column=10).value = f"=SUM(J14:J{civ_row})"
    # sheet.cell(row=civ_row+3, column=12).value = f"=SUM(L14:L{civ_row})"
    # sheet.cell(row=civ_row+3, column=15).value = f"=SUM(O14:O{civ_row})"
    # sheet.cell(row=civ_row+3, column=16).value = f"=SUM(P14:P{civ_row})"
    # sheet.cell(row=civ_row+3, column=17).value = f"=SUM(Q14:Q{civ_row})"

    # logger.info(f"data: {data}")
    average_single_weight = sum([item.get("GrossWeight") for item in data if not item.get("single_weight")]) /  sum([item.get("carton") for item in data if not item.get("single_weight")])
    #添加单箱重量
    sheet.cell(row=civ_row+7, column=15).value = "单箱重量: "
    sheet.cell(row=civ_row+7, column=16).value = f"{round(average_single_weight,2)}"

    sheet.cell(row=civ_row+8, column=15).value = "预估总税金: "
    sheet.cell(row=civ_row+8, column=16).value = f"{data[0].get('estimated_tax_amount','')}"

    # 保存文件
    output_path = f"file/{time.time()}-{data[0]['MasterBillNo']} CI&PL-{totalyugutax}_admin_审核.xlsx"
    wb.save(output_path)
    return output_path 
# def generate_excel_from_template_test(data,totalyugutax,port):
#     apcelllic = License()
#     apcelllic.setLicense('JAVA-Aspose.Excel-24.7/license.xml')
#     template_path = "./file/excel_template/副本清关发票箱单模板 - 0918更新.xlsx"
#     wb = Workbook(template_path)
#     civ_sheet = wb.getWorksheets().get("CIV")
#     pl_sheet = wb.getWorksheets().get("PL")
#     huomian_explaination_sheet = wb.getWorksheets().get("豁免说明")
#     start_row = 13

    

#     def set_cell_value(sheet, row, column, value):
#         cell = sheet.getCells().get(row - 1, column - 1)  # Adjust for zero-based index
        
#         if isinstance(value, str) and value.startswith('='):
#             # 如果 value 是一个公式
#             cell.setFormula(value)
#         else:
#             # 否则，设置为普通值
#             cell.putValue(value)
        
#         style = cell.getStyle()
#         style.setTextWrapped(True)
#         style.setHorizontalAlignment(TextAlignmentType.CENTER)
#         style.setVerticalAlignment(TextAlignmentType.CENTER)
#         cell.setStyle(style)


#         # Auto-adjust row height
#         # Adjust based on value length, e.g., number of lines
#         # num_lines = value.count("\n") + 1
#         # if num_lines > 1:
#         #     sheet.getCells().setRowHeight(row - 1, num_lines * 15)

#     # Fill CIV content
#     set_cell_value(civ_sheet, 1, 1, data[0]["shipper_name"])
#     set_cell_value(civ_sheet, 2, 1, data[0]["shipper_address"])
#     set_cell_value(civ_sheet, 6, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

#     today_minus_5 = datetime.now() - timedelta(days=5)
#     formatted_date = today_minus_5.strftime("%Y%m%d")
#     random_number = random.randint(1000, 9999)
#     result_1 = f"{formatted_date}{random_number}"
#     set_cell_value(civ_sheet, 6, 9, result_1)
#     set_cell_value(civ_sheet, 7, 9, result_1)
#     set_cell_value(civ_sheet, 8, 9, datetime.now().strftime("%Y/%m/%d"))
#     if data[0]["export_country"] == "Vietnam":
#         set_cell_value(civ_sheet, 9, 9, "MADE IN VIETNAM")
#         set_cell_value(civ_sheet, 9, 2, "")
#         set_cell_value(pl_sheet, 8, 9, "MADE IN VIETNAM")
#         set_cell_value(pl_sheet, 8, 2, "")
#     set_cell_value(pl_sheet, 1, 1, data[0]["shipper_name"])
#     set_cell_value(pl_sheet, 2, 1, data[0]["shipper_address"])
#     set_cell_value(pl_sheet, 5, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

#     set_cell_value(pl_sheet, 5, 9, result_1)
#     set_cell_value(pl_sheet, 6, 9, result_1)
#     set_cell_value(pl_sheet, 7, 9, datetime.now().strftime("%Y/%m/%d"))

#     if data[0]["execute_type"] == "Sea":
#         set_cell_value(civ_sheet, 9, 2, "")
#         set_cell_value(civ_sheet, 10, 2, "US BY SEA")
#         set_cell_value(pl_sheet, 8, 2, "")
#         set_cell_value(pl_sheet, 9, 2, "US BY SEA")
#     else:
#         set_cell_value(civ_sheet, 11, 9, port)
#         set_cell_value(pl_sheet, 10, 9, port)

    
#     all_pic_path = []
#     for index, item in enumerate(data):
#         civ_row = 14 + index
#         huomian_row = 5 + index
#         pl_row = start_row + index


#         # Fill CIV
#         set_cell_value(civ_sheet, civ_row, 1, "=ROW()-ROW($A$14)+1")
#         set_cell_value(civ_sheet, civ_row, 2, item["HS_CODE"])
#         set_cell_value(civ_sheet, civ_row, 3, item["DESCRIPTION"])
#         set_cell_value(civ_sheet, civ_row, 4, item["quanity"])
#         set_cell_value(civ_sheet, civ_row, 5, item["danwei"])
#         set_cell_value(civ_sheet, civ_row, 6, item["unit_price"])
#         set_cell_value(civ_sheet, civ_row, 7, round(item["total_price"]))
#         set_cell_value(civ_sheet, civ_row, 8, item["texture"])
#         set_cell_value(civ_sheet, civ_row, 9, item["address_name"])
#         set_cell_value(civ_sheet, civ_row, 10, item["address"])
#         set_cell_value(civ_sheet, civ_row, 11, item["note"])

#         civ_sheet.autoFitRow(civ_row - 1)


#         # Fill PL
#         set_cell_value(pl_sheet, pl_row, 1, "=ROW()-ROW($A$13)+1")

#         set_cell_value(pl_sheet, pl_row, 2, item["HS_CODE"])

#         set_cell_value(pl_sheet, pl_row, 3, item["DESCRIPTION"])
#         set_cell_value(pl_sheet, pl_row, 4, item["quanity"])
#         set_cell_value(pl_sheet, pl_row, 5, item["danwei"])
#         set_cell_value(pl_sheet, pl_row, 6, item["carton"])
#         set_cell_value(pl_sheet, pl_row, 8, item["net_weight"])
#         set_cell_value(pl_sheet, pl_row, 9, item["GrossWeight"])
#         set_cell_value(pl_sheet, pl_row, 10, item["Volume"])
#         pl_sheet.autoFitRow(pl_row - 1)

#         # Fill 豁免说明
#         set_cell_value(huomian_explaination_sheet, huomian_row, 1, item["HS_CODE"])
#         set_cell_value(huomian_explaination_sheet, huomian_row, 2, item["DESCRIPTION"])
#         set_cell_value(huomian_explaination_sheet, huomian_row, 3, item["usage"])
#         set_cell_value(huomian_explaination_sheet, huomian_row, 4, item["note"])
#         set_cell_value(huomian_explaination_sheet, huomian_row, 5, item["note_explaination"])
#         if item["huomian_file_name"]:
#             pic_path = os.path.join("./file/huomian_file/",item["huomian_file_name"])

#             all_pic_path.append({"pic_path":pic_path,"new_name":item['DESCRIPTION']})
#         huomian_explaination_sheet.autoFitRow(huomian_row - 1)

#         if index  == len(data) - 1 :
#             #如果是最后一个循环的数据，则不需要再添加一行了
#             break
#         # 在每个循环结束时增加一行，以避免覆盖
#         # print(civ_sheet.getCells().get(civ_row, 2).getValue())
#         if civ_sheet.getCells().get(civ_row + 1, 2).getValue() == "TOTAL":
#             civ_sheet.getCells().insertRows(civ_row, 1)
#         if pl_sheet.getCells().get(pl_row+ 1, 2).getValue() == "TOTAL":

#             pl_sheet.getCells().insertRows(pl_row, 1)
#         # pl_sheet.getCells().insertRows(pl_row, 1)

#         # civ_sheet.getCells().insertRows(civ_row, 1)
#         huomian_explaination_sheet.getCells().insertRows(huomian_row, 1)

#     civ_sheet.getCells().hideColumn(5)


#     # Save the Excel file
#     output_path = f"file/{time.time()}-{data[0]['MasterBillNo']} CI&PL-{totalyugutax}.xlsx"
#     wb.calculateFormula()
#     wb.save(output_path, SaveFormat.XLSX)
#     print(f"Excel file generated: {output_path}")

#     # 生成 PDF 文件
#     pdf_path = excel2pdf(output_path, 'pdf')

#     # 压缩图片和PDF文件
#     # if all_pic_path:
#     #     zip_path = f"file/{time.time()}-{data[0]['MasterBillNo']} CI&PL-{totalyugutax}.zip"
#     #     with zipfile.ZipFile(zip_path, 'w') as zipf:
#     #         zipf.write(pdf_path, os.path.basename(pdf_path))
#     #         for index, item in enumerate(all_pic_path):
#     #             pic_path = item['pic_path']
#     #             new_pic_name = os.path.join(os.path.dirname(pic_path),item['new_name']+".png")
#     #             zipf.write(pic_path, new_pic_name)
#     #     logger.info(f"zip文件已成功生成: {zip_path}")
#     #     return zip_path
#     logger.info(f"pdf文件已成功生成: {pdf_path}")
#     return pdf_path

# def generate_excel_from_template_canada(data,totalyugutax,currentcy_type="CAD"):
#     apcelllic = License()
#     apcelllic.setLicense('JAVA-Aspose.Excel-24.7/license.xml')
#     template_path = "./file/excel_template/加拿大-清关发票箱单开发模板-0410.xlsx"
#     wb = Workbook(template_path)
#     civ_sheet = wb.getWorksheets().get("CIV")
#     pl_sheet = wb.getWorksheets().get("PL")
   

    

#     def set_cell_value(sheet, row, column, value):
#         cell = sheet.getCells().get(row - 1, column - 1)  # Adjust for zero-based index
        
#         if isinstance(value, str) and value.startswith('='):
#             # 如果 value 是一个公式
#             cell.setFormula(value)
#         else:
#             # 否则，设置为普通值
#             cell.putValue(value)
        
#         style = cell.getStyle()
#         style.setTextWrapped(True)
#         style.setHorizontalAlignment(TextAlignmentType.CENTER)
#         style.setVerticalAlignment(TextAlignmentType.CENTER)
#         cell.setStyle(style)


#         # Auto-adjust row height
#         # Adjust based on value length, e.g., number of lines
#         # num_lines = value.count("\n") + 1
#         # if num_lines > 1:
#         #     sheet.getCells().setRowHeight(row - 1, num_lines * 15)

#     # Fill CIV content
#     # set_cell_value(civ_sheet, 1, 1, data[0]["shipper_name"])
#     # set_cell_value(civ_sheet, 2, 1, data[0]["shipper_address"])
#     # set_cell_value(civ_sheet, 6, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

#     today_minus_5 = datetime.now() - timedelta(days=5)
#     formatted_date = today_minus_5.strftime("%Y%m%d")
#     random_number = random.randint(1000, 9999)
#     result_1 = f"{formatted_date}{random_number}"
#     set_cell_value(civ_sheet, 5, 8, result_1)
#     set_cell_value(civ_sheet, 6, 8, result_1)
#     set_cell_value(civ_sheet, 7, 8, datetime.now().strftime("%Y/%m/%d"))
#     set_cell_value(civ_sheet, 9, 8, currentcy_type)
#     # if data[0]["export_country"] == "Vietnam":
#     #     set_cell_value(civ_sheet, 9, 9, "MADE IN VIETNAM")
#     #     set_cell_value(pl_sheet, 8, 9, "MADE IN VIETNAM")
#     # set_cell_value(pl_sheet, 1, 1, data[0]["shipper_name"])
#     # set_cell_value(pl_sheet, 2, 1, data[0]["shipper_address"])
#     # set_cell_value(pl_sheet, 5, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

#     set_cell_value(pl_sheet, 5, 8, result_1)
#     set_cell_value(pl_sheet, 6, 8, result_1)
#     set_cell_value(pl_sheet, 7, 8, datetime.now().strftime("%Y/%m/%d"))

#     # if data[0]["execute_type"] == "Sea":
#     #     set_cell_value(civ_sheet, 9, 2, "")
#     #     set_cell_value(civ_sheet, 10, 2, "US BY SEA")
#     #     set_cell_value(pl_sheet, 8, 2, "")
#     #     set_cell_value(pl_sheet, 9, 2, "US BY SEA")
    
    
#     all_pic_path = []
#     for index, item in enumerate(data):
#         civ_row = 13 + index
#         pl_row = 13 + index


#         # Fill CIV
#         set_cell_value(civ_sheet, civ_row, 1, "=ROW()-ROW($A$13)+1")
#         set_cell_value(civ_sheet, civ_row, 2, item["HS_CODE"])
#         set_cell_value(civ_sheet, civ_row, 3, item["DESCRIPTION"])
#         set_cell_value(civ_sheet, civ_row, 4, item["quanity"])
#         set_cell_value(civ_sheet, civ_row, 5, item["danwei"])
#         set_cell_value(civ_sheet, civ_row, 6, item["unit_price"])
#         set_cell_value(civ_sheet, civ_row, 7, round(item["total_price"]))
#         set_cell_value(civ_sheet, civ_row, 8, item["texture"])
#         # set_cell_value(civ_sheet, civ_row, 9, item["address_name"])
#         # set_cell_value(civ_sheet, civ_row, 10, item["address"])
#         set_cell_value(civ_sheet, civ_row, 9, item["note"])

#         civ_sheet.autoFitRow(civ_row - 1)


#         # Fill PL
#         set_cell_value(pl_sheet, pl_row, 1, "=ROW()-ROW($A$13)+1")

#         set_cell_value(pl_sheet, pl_row, 2, item["HS_CODE"])

#         set_cell_value(pl_sheet, pl_row, 3, item["DESCRIPTION"])
#         set_cell_value(pl_sheet, pl_row, 4, item["quanity"])
#         set_cell_value(pl_sheet, pl_row, 5, item["danwei"])
#         set_cell_value(pl_sheet, pl_row, 6, item["carton"])
#         set_cell_value(pl_sheet, pl_row, 8, item["net_weight"])
#         set_cell_value(pl_sheet, pl_row, 9, item["GrossWeight"])
#         set_cell_value(pl_sheet, pl_row, 10, item["Volume"])
#         pl_sheet.autoFitRow(pl_row - 1)

#         # Fill 豁免说明
      
#         if item["huomian_file_name"]:
#             pic_path = os.path.join("./file/huomian_file/",item["huomian_file_name"])

#             all_pic_path.append({"pic_path":pic_path,"new_name":item['DESCRIPTION']})

#         if index  == len(data) - 1 :
#             #如果是最后一个循环的数据，则不需要再添加一行了
#             break
#         # 在每个循环结束时增加一行，以避免覆盖
#         # print(civ_sheet.getCells().get(civ_row, 2).getValue())
#         if civ_sheet.getCells().get(civ_row + 1, 2).getValue() == "TOTAL":
#             civ_sheet.getCells().insertRows(civ_row, 1)
#         if pl_sheet.getCells().get(pl_row+ 1, 2).getValue() == "TOTAL":

#             pl_sheet.getCells().insertRows(pl_row, 1)
#         # pl_sheet.getCells().insertRows(pl_row, 1)

#         # civ_sheet.getCells().insertRows(civ_row, 1)

#     civ_sheet.getCells().hideColumn(5)


#     # Save the Excel file
#     output_path = f"file/{time.time()}-{data[0]['MasterBillNo']} CI&PL-{totalyugutax}.xlsx"
#     wb.calculateFormula()
#     wb.save(output_path, SaveFormat.XLSX)
#     print(f"Excel file generated: {output_path}")

#     # 生成 PDF 文件
#     pdf_path = excel2pdf(output_path, 'pdf')

#     # 压缩图片和PDF文件
#     # if all_pic_path:
#     #     zip_path = f"file/{time.time()}-{data[0]['MasterBillNo']} CI&PL-{totalyugutax}.zip"
#     #     with zipfile.ZipFile(zip_path, 'w') as zipf:
#     #         zipf.write(pdf_path, os.path.basename(pdf_path))
#     #         for index, item in enumerate(all_pic_path):
#     #             pic_path = item['pic_path']
#     #             new_pic_name = os.path.join(os.path.dirname(pic_path),item['new_name']+".png")
#     #             zipf.write(pic_path, new_pic_name)
#     #     logger.info(f"zip文件已成功生成: {zip_path}")
#     #     return zip_path
#     logger.info(f"pdf文件已成功生成: {pdf_path}")
#     return pdf_path

# def generate_excel_from_template(data):
#     template_path = r"清关发票箱单模板.xlsx"
#     # 读取模板文件
#     wb = openpyxl.load_workbook(template_path)
#     civ_sheet = wb["CIV"]
#     pl_sheet = wb["PL"]
#     huomian_explaination_sheet = wb['豁免说明']
#     start_row = 13
    

#     def set_cell_value(sheet, row, column, value):
#         cell = sheet.cell(row=row, column=column)
#         # 检查是否为合并单元格，如果是，只在左上角单元格写入值
#         if cell.coordinate in sheet.merged_cells:
#             for merged_range in sheet.merged_cells.ranges:
#                 if cell.coordinate in merged_range:
#                     top_left_cell = merged_range.start_cell
#                     sheet[top_left_cell.coordinate].value = value
#                     sheet[top_left_cell.coordinate].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
#                     break
#         else:
#             cell.value = value
#             cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

#         # 自动调整行高
#         wrap_length = 30  # 根据需要调整换行长度
#         lines = textwrap.wrap(str(value), wrap_length)
#         num_lines = len(lines)
#         # 只有在内容需要换行时才调整行高
#         if num_lines > 1:
#             sheet.row_dimensions[row].height = num_lines * 15

 

#     # 填充 CIV 一次内容
#     set_cell_value(civ_sheet, 1, 1, data[0]["shipper_name"])
#     set_cell_value(civ_sheet, 2, 1, data[0]["shipper_address"])
#     set_cell_value(civ_sheet, 6, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

#     today_minus_5 = datetime.now() - timedelta(days=5)
#     formatted_date = today_minus_5.strftime("%Y%m%d")
#     random_number = random.randint(1000, 9999)
#     result_1 = f"{formatted_date}{random_number}"
#     set_cell_value(civ_sheet, 6, 8, result_1)
#     set_cell_value(civ_sheet, 7, 8, result_1)
#     set_cell_value(civ_sheet, 8, 8, datetime.now().strftime("%Y/%m/%d"))

#     set_cell_value(pl_sheet, 1, 1, data[0]["shipper_name"])
#     set_cell_value(pl_sheet, 2, 1, data[0]["shipper_address"])
#     set_cell_value(pl_sheet, 5, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

#     set_cell_value(pl_sheet, 5, 8, result_1)
#     set_cell_value(pl_sheet, 6, 8, result_1)
#     set_cell_value(pl_sheet, 7, 8, datetime.now().strftime("%Y/%m/%d"))

#     if data[0]["execute_type"] == "Sea":
#         set_cell_value(civ_sheet, 9, 2, "")
#         set_cell_value(civ_sheet, 10, 2, "US BY SEA")
#         set_cell_value(pl_sheet, 8, 2, "")
#         set_cell_value(pl_sheet, 9, 2, "US BY SEA")

#     for index, item in enumerate(data):
#         civ_row = 14 + index
#         huomian_row = 5 + index
#         pl_row = start_row + index

#         # 如果行数超出最大值，在最大行上方插入新行
#         # civ_row = insert_row_if_needed(civ_sheet, civ_row, civ_max_row)
#         # huomian_row = insert_row_if_needed(huomian_explaination_sheet, huomian_row, huomian_max_row)
#         # pl_row = insert_row_if_needed(pl_sheet, pl_row, pl_max_row)

#         # 填充 CIV
#         set_cell_value(civ_sheet, civ_row, 1, item["HS_CODE"])
#         set_cell_value(civ_sheet, civ_row, 2, item["DESCRIPTION"])
#         set_cell_value(civ_sheet, civ_row, 3, item["quanity"])
#         set_cell_value(civ_sheet, civ_row, 4, item["danwei"])
#         set_cell_value(civ_sheet, civ_row, 5, item["unit_price"])
#         set_cell_value(civ_sheet, civ_row, 6, item["total_price"])
#         set_cell_value(civ_sheet, civ_row, 7, item["texture"])
#         set_cell_value(civ_sheet, civ_row, 8, item["address_name"])
#         set_cell_value(civ_sheet, civ_row, 9, item["address"])
#         set_cell_value(civ_sheet, civ_row, 10, item["note"])

#         # 填充 PL
#         set_cell_value(pl_sheet, pl_row, 2, item["DESCRIPTION"])
#         set_cell_value(pl_sheet, pl_row, 3, item["quanity"])
#         set_cell_value(pl_sheet, pl_row, 4, item["danwei"])
#         set_cell_value(pl_sheet, pl_row, 5, item["carton"])
#         set_cell_value(pl_sheet, pl_row, 7, item["net_weight"])
#         set_cell_value(pl_sheet, pl_row, 8, item["GrossWeight"])
#         set_cell_value(pl_sheet, pl_row, 9, item["Volume"])

#         # 填充豁免说明
#         set_cell_value(huomian_explaination_sheet, huomian_row, 1, item["HS_CODE"])
#         set_cell_value(huomian_explaination_sheet, huomian_row, 2, item["DESCRIPTION"])
#         set_cell_value(huomian_explaination_sheet, huomian_row, 3, item["note"])
#         set_cell_value(huomian_explaination_sheet, huomian_row, 4, item["note_explaination"])

#     # 保存新的 Excel 文件
#     output_path = f"file/{data[0]['MasterBillNo']} CI&PL.xlsx"
#     wb.save(output_path)
#     logger.info(f"excel文件已成功生成: {output_path}")

#     # 生成 PDF 文件
#     pdf_path = excel2pdf(output_path, 'pdf')
#     logger.info(f"pdf文件已成功生成: {pdf_path}")
    
#     return pdf_path

# def excel2pdf(excel_path: str, pdf_save_path: str = None) -> str:
#     # 加载License文件
#     apcelllic = License()
#     apcelllic.setLicense('JAVA-Aspose.Excel-24.7/license.xml')

#     # 打开Excel文件
#     wb = Workbook(excel_path)
#     # 删除名称为 "Evaluation Warning" 的工作表（如果存在）
#     sheets = wb.getWorksheets()
#     eval_warning_sheet = sheets.get("Evaluation Warning")
#     if eval_warning_sheet is not None:
#         sheets.removeAt("Evaluation Warning")
#     # 配置PDF保存选项
#     saveOption = PdfSaveOptions()
    
#     # 确保每个工作表单独保存为一个PDF页面
#     # saveOption.setOnePagePerSheet(True)  # 如果为True，将整个工作表压缩到一个PDF页面上
#     saveOption.setAllColumnsInOnePagePerSheet(True) #所有列在一页，但是可能行在多页

#     # 计算公式
#     saveOption.setCalculateFormula(True)  # 计算公式并将其值保存在PDF中

#     # 设置字体相关选项
#     saveOption.setCheckWorkbookDefaultFont(True)  # 检查工作簿的默认字体，以避免出现方块字符
#     saveOption.setCheckFontCompatibility(True)  # 检查每个字符的字体兼容性
#     saveOption.setDefaultFont("Arial")  # 设置默认字体（如果未设置正确的字体）

#     # 设置图像处理
#     saveOption.setImageResample(220, 85)  # 设置图像的PPI和JPEG质量，减少PDF文件大小

#     # 设置其他相关选项
#     saveOption.setEmbedStandardWindowsFonts(True)  # 嵌入标准的Windows字体
#     saveOption.setClearData(False)  # 在保存后不清除工作簿的数据
#     saveOption.setCompliance(0)  # 设置PDF标准合规级别，如需要合规的PDF/A等格式
#     saveOption.setDisplayDocTitle(True)  # 在PDF窗口的标题栏显示文档标题

#     # 如果没有指定保存路径，则使用与 Excel 文件相同的路径
#     if pdf_save_path is None:
#         pdf_save_path = os.path.dirname(excel_path)
    
#     # 获取Excel文件的文件名（不含扩展名）
#     excel_name = os.path.splitext(os.path.basename(excel_path))[0]
    
#     # 设置PDF文件的完整保存路径
#     pdf_file = os.path.join(pdf_save_path, f"{excel_name}.pdf")

#     # 保存为PDF
#     wb.save(pdf_file, saveOption)
    
#     return pdf_file
# 自定义处理器，用于在日志记录时发送邮件
def email_handler(message, receiver_email):
    record = message.record
    if record["level"].name == "ALERT":
        subject = record["message"]
        body = f"错误如下:\n\{record['message']}"
        send_email(subject=subject, receiver_email=receiver_email, body=body)


def create_email_handler(to_addrs):
    return lambda message: email_handler(message, to_addrs)
# def shenzhen_customes_pdf_gennerate(data, filter_data):
#     apcelllic = License()
#     apcelllic.setLicense("JAVA-Aspose.Excel-24.7/license.xml")
#     template_path = "HAWB模板-空+海_测试新版.xls"
#     wb = Workbook(template_path)
#     shenzhn_sheet = wb.getWorksheets().get("S#-SZ-customs")

#     def set_cell_value(sheet, row, column, value):
#         if value is None:
#             value = ""
#         cell = sheet.getCells().get(row - 1, column - 1)  # Adjust for zero-based index

#         if isinstance(value, str) and value.startswith("="):
#             # 如果 value 是一个公式
#             cell.setFormula(value)
#         else:
#             # 否则，设置为普通值
#             cell.putValue(value)

#         # style = cell.getStyle()
#         # style.setTextWrapped(True)
#         # style.setHorizontalAlignment(TextAlignmentType.CENTER)
#         # style.setVerticalAlignment(TextAlignmentType.CENTER)
#         # cell.setStyle(style)

#     set_cell_value(shenzhn_sheet, 5, 4, data["shipper_name"])
#     set_cell_value(shenzhn_sheet, 5, 17, data["master_bill_no"])

#     set_cell_value(shenzhn_sheet, 9, 4, data["receiver_name"])
#     set_cell_value(shenzhn_sheet, 16, 4, data["receiver_name"])
#     set_cell_value(shenzhn_sheet, 26, 10, data["total_boxes"])
#     set_cell_value(shenzhn_sheet, 26, 15, data["all_english_name"])
#     set_cell_value(shenzhn_sheet, 26, 22, str(data["gross_weight"]))
#     set_cell_value(shenzhn_sheet, 26, 25, str(data["volume"]))

#     set_cell_value(shenzhn_sheet, 48, 14, data["gross_weight"])
#     set_cell_value(shenzhn_sheet, 48, 19, str(data["volume"])+"cbm")
#     if filter_data["hblno"]:
#         set_cell_value(shenzhn_sheet, 5, 24, filter_data["hblno"])
#     else:
#         set_cell_value(shenzhn_sheet, 5, 24, "系统未录入")

#     set_cell_value(shenzhn_sheet, 19, 11, filter_data["startland"])
#     set_cell_value(shenzhn_sheet, 21, 11, filter_data["startland"])
#     set_cell_value(shenzhn_sheet, 23, 4, filter_data["destination"])
#     set_cell_value(shenzhn_sheet, 23, 11, filter_data["destination"])
#     # 航次
#     set_cell_value(shenzhn_sheet, 21, 4, filter_data["flight"])

#     # 柜号
#     set_cell_value(shenzhn_sheet, 48, 5, filter_data["cabinetNo"])
#     # 封条号
#     set_cell_value(shenzhn_sheet, 48, 6, filter_data["sealno"])
#     # 柜型
#     set_cell_value(shenzhn_sheet, 48, 8, filter_data["cabinettype"])
#     if filter_data["ATD"]:
#         try:
#             atd_datetime = datetime.strptime(filter_data["ATD"], "%Y-%m-%d %H:%M")
            
#         except ValueError:
#             # 如果解析失败，尝试解析不包含秒的格式
#             atd_datetime = datetime.strptime(filter_data["ATD"], "%Y-%m-%d %H:%M:%S")

#         # 保留到年月日
#         filter_data["ATD"] = atd_datetime.strftime("%Y-%m-%d")
#         set_cell_value(shenzhn_sheet, 50, 13, filter_data["ATD"])
#     else:
#         set_cell_value(shenzhn_sheet, 50, 13, "系统未录入")


#     ids = []
#     for i in wb.getWorksheets():
#         origin_sheetname = i.getName()
#         if origin_sheetname == "S#-SZ-customs":
#             ids.append(i.getIndex())
#     new_SheetSet = SheetSet(ids)

#     # 配置PDF保存选项
#     saveOption = PdfSaveOptions()
#     saveOption.setSheetSet(new_SheetSet)
#     # 确保每个工作表单独保存为一个PDF页面
#     saveOption.setOnePagePerSheet(True)  # 如果为True，将整个工作表压缩到一个PDF页面上

#     # 计算公式
#     saveOption.setCalculateFormula(True)  # 计算公式并将其值保存在PDF中

#     # 设置字体相关选项
#     saveOption.setCheckWorkbookDefaultFont(
#         True
#     )  # 检查工作簿的默认字体，以避免出现方块字符
#     saveOption.setCheckFontCompatibility(True)  # 检查每个字符的字体兼容性
#     saveOption.setDefaultFont("Arial")  # 设置默认字体（如果未设置正确的字体）

#     # 设置图像处理
#     saveOption.setImageResample(220, 85)  # 设置图像的PPI和JPEG质量，减少PDF文件大小

#     # 设置其他相关选项
#     saveOption.setEmbedStandardWindowsFonts(True)  # 嵌入标准的Windows字体
#     saveOption.setClearData(False)  # 在保存后不清除工作簿的数据
#     saveOption.setCompliance(0)  # 设置PDF标准合规级别，如需要合规的PDF/A等格式
#     saveOption.setDisplayDocTitle(True)  # 在PDF窗口的标题栏显示文档标题

#     # 设置PDF文件的完整保存路径
#     totalyugutax = data["other_data"]["totalyugutax"]

#     pdf_file = f"./pdf/customs/{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}-{data['master_bill_no']}-{totalyugutax}.pdf"

#     # 保存为PDF
#     wb.save(pdf_file, saveOption)
#     return pdf_file

def get_session():
    try:
        return Session(pool_engine)
    except Exception as e:
        logger.error(f"Database connection error: {e}")
        raise





def format_excel_sheet(sheet, headers):
    """
    格式化指定的 Excel 工作表，包括：
        - 微软雅黑 10号字体
        - 首行加粗居中
        - 添加筛选
        - 冻结首行
        - 设置列宽自适应

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): 要格式化的 openpyxl 工作表对象。
        headers (list): 列标题列表。

    Returns:
        openpyxl.worksheet.worksheet.Worksheet: 格式化后的 openpyxl 工作表对象 (返回修改后的 sheet)。
    """

    sheet.append(headers)

    # 1. 设置字体和对齐方式
    font = Font(name="微软雅黑", size=10)
    bold_font = Font(name="微软雅黑", size=10, bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 添加自动换行

    # 遍历所有单元格，设置字体
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    # 设置首行字体加粗
    for cell in sheet[1]:
        cell.font = bold_font
        cell.alignment = alignment

    # 2. 添加筛选器
    sheet.auto_filter.ref = "A1:" + get_column_letter(len(headers)) + str(sheet.max_row)

    # 3. 冻结首行
    sheet.freeze_panes = "A2"

    # 4. 设置列宽自适应 (非常重要，但效率可能较低，根据实际情况调整)
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in sheet[column_letter]:
            try:  # 处理单元格内容可能不是字符串的情况
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2  # 适当增加一些宽度
        sheet.column_dimensions[column_letter].width = adjusted_width

    return sheet
def output_custom_clear_history_log(
        start_date:str=None,
        end_date:str=None,
        filename:str=None,
        port:str=None,
        packing_type:str=None,
        good_type:str=None,
        abnormal:str=None,
        remarks:str=None,
        convey_type:str=None,
        id_list:list=None,

):
    try:
        with get_db() as db:
            # 查询所有收货人，发货人
            consignees = list(db.consignees.find())
            query = {}
            
            # 如果有id_list,优先使用id查询
            if id_list:
                query['_id'] = {"$in": [ObjectId(id) for id in id_list]}
            else:
                # 如果没有id_list,使用时间范围和其他条件查询
                query["generation_time"] = {
                    "$gte": datetime.strptime(start_date, "%Y-%m-%d"), 
                    "$lte": datetime.strptime(end_date, "%Y-%m-%d")
                }
                
                if filename:
                    query["filename"] = {"$regex": f".*{filename}.*", "$options": "i"}
                if port:
                    query["port"] = port
                if packing_type:
                    query["packing_type"] = packing_type
                if good_type:
                    query["good_type"] = good_type
                if abnormal:
                    query["abnormal"] = {"$regex": f".*{abnormal}.*", "$options": "i"}
                if remarks:
                    query["remarks"] = {"$regex": f".*{remarks}.*", "$options": "i"}
                if convey_type:
                    query["convey_type"] = convey_type

            summary_logs = list(
                db.custom_clear_history_summary.find(query)
            )
            # logger.info(f"summary_logs:{summary_logs}")

            # 1 .从汇总日志中获取明细数据
            # 2. 从汇总数据中获取分组（总单列表，大货运单）
            # 3. 如果汇总数据的详情数据没有box_num,则从excel中读取
            detail_logs = []
            dahuo_group_summary_logs = []
            zongdan_group_detail_logs = []
            data_df = None
            for summary in summary_logs:
                if isinstance(summary.get('remarks'), str) and '/' in summary['remarks']:
                    try:
                        summary["real_tax_amount"] = float(summary['remarks'].split('/')[0])
                    except ValueError:
                        summary["real_tax_amount"] = None  # Handle cases where conversion fails
                else:
                    summary["real_tax_amount"] = summary.get('remarks')

                summary["主单号"] = '-'.join(summary['filename'].split('-')[1:-1]).replace('CI&PL','').strip()
                summary['rate'] = float(summary.get('rate',7.3))
                #3
                detail_data_box_num = {}
                if summary.get("details") and summary["details"] and not summary["details"][0].get("gross_weight") :
                    file_path = os.path.join("./file/", summary['filename'].replace('pdf','xlsx'))
                    try:
                        logger.info(f"file_path:{file_path}") 
                        pl_df = pd.read_excel(file_path, sheet_name="PL", skiprows=11)
                        # 确保 'HS CODE' 和 'CARTON' 列存在
                        if 'HS CODE' not in pl_df.columns or 'CARTON' not in pl_df.columns :
                            print("Error: 'HS CODE' or 'CARTON' column not found in the Excel sheet.")
                            # return None # 不应该直接返回None，而是继续处理其他数据
                            continue
                        civ_df = pd.read_excel(file_path, sheet_name="CIV", skiprows=12)
                        civ_df = civ_df.dropna(subset=['HS CODE'])
                        civ_df['TOTAL PRICE (USD)'] = pd.to_numeric(civ_df['TOTAL PRICE (USD)'], errors='coerce')

                        total_price_sum = civ_df['TOTAL PRICE (USD)'].sum()
                        summary['total_price_sum'] = total_price_sum
                        # 移除包含NaN值的行，避免后续转换出错
                        pl_df = pl_df.dropna(subset=['HS CODE', 'CARTON'])
 
                        
                        # 将HS CODE列转换为字符串类型，避免出现数字格式问题
                        pl_df['HS CODE'] = pl_df['HS CODE'].astype(str)

                        # 将DataFrame转换为字典
                        detail_data_box_num = dict(zip(pl_df['HS CODE'], pl_df['CARTON']))

                        # logger.info(f"pl_df:{pl_df}")

                        #pl_df和civ_df合并（HS CODE 和 QUANTITY）
                        data_df = pl_df.merge(civ_df, on=['HS CODE'], how='left')
                        # data_df = pd.concat([pl_df, civ_df], axis=1)
                        # logger.info(f"data_df columns:{data_df.columns}")

                    except FileNotFoundError:
                        print(f"文件未找到: {summary['filename']}")
                    except Exception as e:
                        print(f"读取Excel文件出错: {e}")
                
                #1
                if summary.get("details") and isinstance(summary["details"], list) and summary.get('remarks') != '删除':  # 确保summary中有detail字段
                    for detail in summary["details"]:
                        detail["_id"] = summary["_id"]  # 添加summary_id
                        detail["generation_time"] = summary["generation_time"]
                        detail["rate"] = summary["rate"]
                        if detail_data_box_num and detail.get('hs_code') in detail_data_box_num:
                             detail['box_nums'] = detail_data_box_num.get(detail['hs_code'])
                        # else:
                        #     detail['box_nums'] = None

                        if data_df is not None:
                            try:
                                detail['unit_price'] = data_df.loc[data_df['HS CODE'] == detail['hs_code'], 'UNIT PRICE (USD)'].values[0]
                                detail['quantity'] = data_df.loc[data_df['HS CODE'] == detail['hs_code'], 'QUANTITY_x'].values[0]
                                detail['piece_per_carton'] = detail['quantity']/detail['box_nums']
                                detail['total_price'] = data_df.loc[data_df['HS CODE'] == detail['hs_code'], 'TOTAL PRICE (USD)'].values[0]
                                detail['net_weight'] = data_df.loc[data_df['HS CODE'] == detail['hs_code'], 'N.W（KGS)'].values[0]
                                detail['gross_weight'] = data_df.loc[data_df['HS CODE'] == detail['hs_code'], 'G.W（KGS)'].values[0]
                                detail['volume'] = data_df.loc[data_df['HS CODE'] == detail['hs_code'], 'Measurement\n(M3)'].values[0]
                                detail['a_box_weight'] = detail['gross_weight']/detail['box_nums']
                                detail['a_piece_weight'] = detail['net_weight']/detail['quantity']
                            except Exception as e:
                                logger.error(f"Error processing detail data: {traceback.format_exc()}")
                                continue
                        else:
                            detail['unit_price'] = float(detail.get("single_price",0.0))
                            detail['quantity'] = float(detail.get("packing",0.0)) * float(detail.get("box_nums",0.0))
                            detail['piece_per_carton'] = float(detail.get("packing",0.0))
                            detail['total_price'] = float(detail['unit_price']) * float(detail['quantity'])
                            detail['net_weight'] = float(detail.get("net_weight",0.0))
                            detail['gross_weight'] = float(detail.get("gross_weight",0.0))
                            detail['volume'] = float(detail.get("volume",0.0))
                            detail['a_box_weight'] = float(detail['gross_weight'])/float(detail.get("box_nums",0.0))
                            detail['a_piece_weight'] = float(detail['net_weight'])/float(detail['quantity'])
                            
                        detail_logs.append(detail)


                #2
                if  summary["主单号"].startswith("A2") or summary["主单号"].startswith("S2"):
                    dahuo_group_summary_logs.append(summary["主单号"])
                else:
                    zongdan_group_detail_logs.append(summary["主单号"])



            node_path = find_playwright_node_path()
            morelink_client = MoreLinkClient(node_path)
            #morelink查询大货运单
            start_date_ml = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=30)).strftime("%Y-%m-%d")
            ml_dahuo_group_detail_logs = morelink_client.dahuodingdan_worknum_search_httpx(signtype='多单号',numberno=','.join(dahuo_group_summary_logs),start_date=start_date_ml,end_date=end_date)
            ml_zongdan_group_detail_logs = morelink_client.zongdan_api_httpx(start_date=start_date_ml,end_date=end_date)
            # 创建 Excel 文件
            wb = Openpyxl_Workbook()
            ws_summary = wb.active
            ws_summary.title = "历史汇总"
            ws_detail = wb.create_sheet(title="历史明细")

            # 写入汇总日志数据
            summary_headers = [
                "空运目的港",
                "海运装箱",
                "发货人",
                "收货人",
                "生成日期",
                "更新时间",
                "M#",
                "ETD",
                "运输方式",
               
                '类别',
                '整票货值',
                "预估整票税金",
                "整票重量",
                "整票体积",
                "整票箱数",
                "预估税金CNY/KG",
                "汇率",
                "核算",
                "真实税金",
                "总税金差额",
                "真实税金CNY/KG",
                "CNY/KG差异",
                "备注",
                "异常", 
                '月份',
                '操作',
                "说明",
                "ID",
                "creator"
            ]
            # ws_summary.append(summary_headers)
            ws_summary = format_excel_sheet(ws_summary, summary_headers)
            all_summary_id = []
            ETD_Dict = {}
            for log in summary_logs:
                all_summary_id.append(str(log["_id"]))
                # 查找发货人和收货人的中文名称
                shipper_cn = next((consignee["中文"] for consignee in consignees if consignee["发货人"] == log["shipper"]), None)
                consignee_cn = next((consignee["中文"] for consignee in consignees if consignee["发货人"] == log["consignee"]), None)
                if str(log.get("主单号", "")).startswith("A2") or str(log.get("主单号", "")).startswith("S2"):
                    ETD = [d['z_etd'] for d in ml_dahuo_group_detail_logs if d['operNo'] == log.get('主单号', '')]
                else:
                    ETD = [d['etd'] for d in ml_zongdan_group_detail_logs if d['billno'] == log.get('主单号', '')]

                Month = ""
                ETD = ETD[0] if ETD else "Error"
                if ETD != "Error" and isinstance(ETD, str):
                    try:
                        Month = datetime.strptime(ETD, "%Y-%m-%d").strftime("%m")
                    except ValueError:
                        Month = ""
                else:
                    Month = ""
                ETD_Dict[log['主单号']] = [ETD,Month]
                if log.get('port'):
                    zongdan_port = [d['destination'] for d in ml_zongdan_group_detail_logs if d['billno'] == log['主单号']]
                    if zongdan_port:
                        log['port'] = zongdan_port[0]
                # logger.info(f"ETD:{log}")
                ws_summary.append(
                    [
                        log.get("port", ""),
                        log.get("packing_type", ""),
                        shipper_cn,  # 使用查找后的中文名称
                        consignee_cn,  # 使用查找后的中文名称
                        log.get("generation_time", datetime.now()).strftime("%Y-%m-%d"),
                        log.get("latest_update_time", datetime.now()).strftime("%Y-%m-%d %H:%M:%S"),
                        log.get('主单号', ''),
                        ETD,
                        "海运" if log.get('packing_type') else ("空运" if log.get('port') else ""),
                        log.get('good_type',''),
                        log.get('total_price_sum',0),

                        log.get("estimated_tax_amount", 0),
                        log.get("gross_weight_kg", 0),
                        log.get("volume_cbm", 0),
                        log.get("total_boxes", 0),
                        log.get("estimated_tax_rate_cny_per_kg", 0),
                        log.get('rate', 0),
                        round(log.get('estimated_tax_amount',0) * float(log.get('rate',0)) / log.get('gross_weight_kg',1), 2) if (log.get('gross_weight_kg') and isinstance(log.get('rate'), (int, float))) else None,
                        log.get('real_tax_amount', 0),
                        (log.get('real_tax_amount',0) - log.get('estimated_tax_amount',0)) if (isinstance(log.get('real_tax_amount'), (int, float)) and isinstance(log.get('estimated_tax_amount'), (int, float))) else None,
                        (round(log.get('real_tax_amount',0) * log.get('rate',0) / log.get('gross_weight_kg',1),2)) if (isinstance(log.get('real_tax_amount'), (int, float)) and log.get('gross_weight_kg')) else None,
                        (log.get('real_tax_amount',0) - log.get('estimated_tax_amount',0)) if (isinstance(log.get('real_tax_amount'), (int, float)) and isinstance(log.get('estimated_tax_amount'), (int, float))) else None,
                        log.get("remarks", ""),
                        log.get("abnormal", ""),
                        Month,
                        "已清关" if log.get("remarks", "") and "未" not in log.get("remarks", "") else "未清关",
                        "",
                        str(log.get("_id", "")),
                        log.get("user_id", "admin")
                    ]
                )

            # 写入明细日志数据
            detail_headers = [
                "HS Code",
                '类型',
                "中文品名",
                "运输方式",
                "主单号",
                "ETD",
                "月份",
                "汇率",
                "生成日期",
                "总税率",
                "豁免代码",
                "单价",
                "数量",
                '件/箱',
                "箱数",

                "总价",
                "净重",
                "毛重",
                "体积",
                '单箱重量',
                '单件重量', 
                "预估税金USD",
                "预估税金CNY/KG",
                "FZ",
                "Summary Log ID",
                "说明"
            ]
            # ws_detail.append(detail_headers)
            ws_detail = format_excel_sheet(ws_detail, detail_headers)
            for log in detail_logs:
                if str(log.get("_id", "")) not in all_summary_id:
                    continue
                
                ws_detail.append(
                    [
                        log.get("hs_code", ""),
                        log.get('category', ''),
                        log.get("chinese_name", ""),
                        "空运" if log.get("transport_mode") == 'Air' else "海运",
                        log.get("master_bill_number", ""),
                        ETD_Dict.get(log.get("master_bill_number", ""), [None, None])[0],
                        ETD_Dict.get(log.get("master_bill_number", ""), [None, None])[1],
                        log.get("rate", 0),
                        log.get("generation_time", datetime.now()).strftime("%Y-%m-%d"),
                        log.get("total_tax_rate", 0),
                        log.get("exemption_code", ""),
                        log.get('unit_price',0),
                        log.get("quantity", 0),
                        log.get("piece_per_carton", 0),
                        log.get('box_nums',0),

                        log.get("total_price", 0),
                        log.get("net_weight", 0),
                        log.get("gross_weight", 0),
                        log.get("volume", 0),
                        round(log.get("a_box_weight", 0),1),
                        round(log.get("a_piece_weight", 0),3),
                        int(log.get('unit_price',0)*log.get('piece_per_carton',0)*log.get('quantity',0))*log.get('total_tax_rate',0),
                        round(int(log.get('unit_price',0)*log.get('piece_per_carton',0)*log.get('quantity',0))*log.get('total_tax_rate',0)*log.get('rate',0)/log.get('net_weight') if log.get('net_weight') else 0,2),
                        "1",
                        str(log.get("_id", "")),
                        
                        ""
                    ]
                )
            #单箱hs这个sheet 添加数据
            # 从历史明细中提取数据,去重后写入单箱hs sheet
            unique_hs_data = {}
            for log in detail_logs:
                if str(log.get("_id", "")) not in all_summary_id:
                    continue
                    
                key = (
                    log.get("hs_code", ""),  # HS CODE
                    log.get("category", ""),  # 类型
                    log.get("chinese_name", ""),  # 中文品名
                    "空运" if log.get("transport_mode") == 'Air' else "海运",  # 运输方式
                    log.get("total_tax_rate", 0),  # 总税率
                    log.get("unit_price", 0),  # 单价
                    log.get("piece_per_carton", 0),  # 件/箱
                    round(log.get('unit_price',0)*log.get('piece_per_carton',0)*(log.get('total_tax_rate',0)+0.003464),2)# 单箱税金
                )
                
                if key not in unique_hs_data:
                    unique_hs_data[key] = True
                    
            # 写入单箱hs sheet
            hs_headers = ["HS CODE", "类型", "中文品名", "运输方式", "总税率", "单价", "件/箱","单箱税金"]
            ws_hs = format_excel_sheet(wb.create_sheet("单箱HS"), hs_headers)
            
            for data in unique_hs_data.keys():
                ws_hs.append(list(data))
            # 保存 Excel 文件到本地
            file_path = f"file/output_log/custom_clear_history_log_{datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            wb.save(file_path)
            # 微软雅黑 10号   首行加粗 居中+筛选+冻结首行  
           

            return file_path

    except Exception as e:
        logger.error(f"Error generating output log: {traceback.format_exc()}")
        print(traceback.format_exc())
        
        raise HTTPException(status_code=500, detail="Internal server error")


def extract_zip_codes_from_excel(excel_path):
    """
    从Excel文件中提取邮政编码
    
    Args:
        excel_path: Excel文件路径
    
    Returns:
        dict: 以sheet名为键，邮政编码列表为值的字典
    """
    # 检查是否存在缓存文件
    cache_file = excel_path + '.json'
    if os.path.exists(cache_file):
        print(f"从缓存文件加载邮政编码数据: {cache_file}")
        with open(cache_file, 'r') as f:
            return json.load(f)
            
    print(f"从Excel文件提取邮政编码数据: {excel_path}")
    
    # 存储结果的字典
    zip_codes_by_sheet = {}
    
    # 读取Excel文件的所有sheet,将所有列作为字符串读取
    df = pd.read_excel(excel_path, sheet_name=None, dtype=str)
    
    # 遍历每个sheet
    for sheet_name, sheet_data in df.items():
        zip_codes = []
        
        # 遍历sheet中的所有列
        for column in sheet_data.columns:
            # 提取该列中的4位或5位数字
            codes = sheet_data[column].str.findall(r'\b\d{4,5}\b')
            # 展平列表并添加到结果中,4位数字前面补0
            for code_list in codes:
                if isinstance(code_list, list):
                    for code in code_list:
                        if len(code) == 4:
                            zip_codes.append('0' + code)
                        else:
                            zip_codes.append(code)
            
        # 去重并存储
        zip_codes_by_sheet[sheet_name] = list(set(zip_codes))
    
    # 将结果保存到缓存文件
    with open(cache_file, 'w') as f:
        json.dump(zip_codes_by_sheet, f)
    return zip_codes_by_sheet



def extract_zip_codes_from_pdf(pdf_path):
    """
    从PDF文件中提取邮政编码
    
    Args:
        pdf_path: PDF文件路径
    
    Returns:
        dict: 以标识为键，邮政编码列表为值的字典
    """
    # 检查是否存在缓存文件
    cache_file = pdf_path + '.json'
    if os.path.exists(cache_file):
        print(f"从缓存文件加载邮政编码数据: {cache_file}")
        with open(cache_file, 'r') as f:
            return json.load(f)
    
    print(f"从PDF文件提取邮政编码数据: {pdf_path}")
    # 打开PDF文件
    pdf_file = open(pdf_path, 'rb')
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    
    # 存储结果的字典
    zip_codes_by_category = {}
    current_category = None
    
    # 遍历PDF的每一页
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        text = page.extract_text()
        
        # 检查页面是否包含标识
        category_match = re.search(r'Deliver y Area Surcharge ZIP codes:\s*(.+?)(?:\n|$)', text)
        
        if category_match:
            # 提取类别名称并清理
            category = category_match.group(1).strip()
            
            # 处理续页情况 (cont.)
            if '(cont.)' in category:
                # 移除(cont.)部分，使用基本类别名称
                base_category = category.replace('(cont.)', '').strip()
                current_category = base_category
            else:
                current_category = category
            
            # 如果是新类别，初始化列表
            if current_category not in zip_codes_by_category:
                zip_codes_by_category[current_category] = []
            
            # 提取邮政编码
            # 使用正则表达式匹配邮政编码格式
            zip_codes = re.findall(r'\b\d{5}(?:-\d{5})?\b', text)
            
            # 将提取的邮政编码添加到当前类别
            zip_codes_by_category[current_category].extend(zip_codes)
    
    # 关闭PDF文件
    pdf_file.close()
    
    # 将结果保存到缓存文件
    with open(cache_file, 'w') as f:
        json.dump(zip_codes_by_category, f)
    
    return zip_codes_by_category

def is_zip_in_range(zip_code, zip_range):
    """
    判断邮政编码是否在指定范围内
    
    Args:
        zip_code: 要检查的邮政编码
        zip_range: 邮政编码范围（单个编码或范围如'01032-01033'）
    
    Returns:
        bool: 如果在范围内返回True，否则返回False
    """
    if '-' in zip_range:
        start, end = zip_range.split('-')
        return start <= zip_code <= end
    else:
        return zip_code == zip_range

def fedex_process_excel_with_zip_codes(input_data, pdf_path=None,excel_path=None):
    """
    处理Excel文件或文本数据，判断每个邮政编码属于哪个分组
    
    Args:
        input_data: Excel文件或包含邮编的文本字符串
        pdf_path: PDF文件路径
    
    Returns:
        DataFrame或dict: 如果输入是Excel则返回DataFrame,如果是文本则返回dict
    """
    # 从PDF中提取邮政编码分组
    if pdf_path:
        zip_codes_by_category = extract_zip_codes_from_pdf(pdf_path)
    # 从excel中获取分组
    if excel_path:
        zip_codes_by_category = extract_zip_codes_from_excel(excel_path)
    # 判断输入是Excel还是文本
    if isinstance(input_data, (str)) and not str(input_data).endswith(('.xlsx', '.xls')):
        # 处理文本输入
        # 统一分隔符为英文逗号
        text = input_data.replace('，', ',').replace(' ', ',')
        zip_codes = [code.strip() for code in text.split(',') if code.strip()]
        
        result = []
        for zip_code in zip_codes:
            # 检查邮编长度
            if len(str(zip_code)) != 5:
                result.append({'type':'fedex','zip_code': zip_code, 'property': '邮编错误,不足五位'})
                continue
                
            # 检查该邮政编码属于哪个分组
            property_found = False
            for category, zip_ranges in zip_codes_by_category.items():
                for zip_range in zip_ranges:
                    if is_zip_in_range(zip_code, zip_range):
                        result.append({'type':'fedex','zip_code': zip_code, 'property': category})
                        property_found = True
                        break
                if property_found:  # 如果已找到分组，跳出循环
                    break
            if not property_found:
                result.append({'type':'fedex','zip_code': zip_code, 'property': 'Unknown'})
                
        return result
        
    else:
        # 处理Excel文件
        df = pd.read_excel(input_data, dtype=str)
        
        # 确保第一列是邮政编码
        zip_column = df.columns[0]
        # 添加新列用于存储分组信息
        df['property'] = ''
        
        # 遍历Excel中的每个邮政编码
        for index, row in df.iterrows():
            zip_code = str(row[zip_column])
            print(zip_code)
            
            # 检查邮编长度
            if len(zip_code) != 5:
                df.at[index, 'property'] = '邮编错误,不足五位'
                continue
                
            # 检查该邮政编码属于哪个分组
            property_found = False
            for category, zip_ranges in zip_codes_by_category.items():
                for zip_range in zip_ranges:
                    if is_zip_in_range(zip_code, zip_range):
                        df.at[index, 'property'] = category
                        property_found = True
                        break
                if property_found:  # 如果已找到分组，跳出循环
                    break
            if not property_found:
                df.at[index, 'property'] = 'Unknown'
        
        return df


def ups_process_excel_with_zip_codes(input_data):
    # 获取property定义Excel文件路径
    property_excel_path = os.path.join(os.getcwd(), 'file', 'remoteaddresscheck', 'area-surcharge-zips-us-en.xlsx')
    
    # 判断输入是Excel还是文本
    if isinstance(input_data, (str, io.BytesIO)) and not str(input_data).endswith(('.xlsx', '.xls')):
        # 处理文本输入
        # 统一分隔符为英文逗号
        text = input_data.replace('，', ',').replace(' ', ',')
        zip_codes = [code.strip() for code in text.split(',') if code.strip()]
        
        # 读取property定义Excel中的所有sheet
        xl = pd.ExcelFile(property_excel_path)
        code_property_map = {}
        
        # 遍历每个sheet获取code和property的对应关系
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(property_excel_path, sheet_name=sheet_name)
            #将数字字符串转换为数字
            data = []
            # 遍历每一列
            for col in df.columns:
                for cell in df[col].dropna():
                    cell = str(cell).zfill(5)
                    if isinstance(str(cell), str):
                        # 使用正则表达式提取数字
                        codes = re.findall(r'\b\d+\b', cell)
                        for code in codes:
                            if code == '00000':
                                continue
                            data.append(code)
            
                code_property_map[sheet_name] = data
            
        # 构建返回的json结果
        result = []
        for zip_code in zip_codes:
            # 检查邮编长度
            if len(str(zip_code)) != 5:
                result.append({'type':'ups','zip_code': zip_code, 'property': '邮编错误,不足五位'})
                continue
                
            property_found = False
            for property_name, codes in code_property_map.items():
                if zip_code in codes:
                    result.append({'type':'ups','zip_code': zip_code, 'property': property_name})
                    property_found = True
                    break
            if not property_found:
                result.append({'type':'ups','zip_code': zip_code, 'property': 'Unknown'})
                
        return result
        
    else:
        # 处理Excel文件
        excel_file = io.BytesIO(input_data)
        input_df = pd.read_excel(excel_file)
        
        if not input_df.empty:
            first_column_name = input_df.columns[0]
            
        # 读取property定义Excel
        xl = pd.ExcelFile(property_excel_path)
        code_property_map = {}
        
        # 遍历每个sheet获取code和property的对应关系
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(property_excel_path, sheet_name=sheet_name)
            data = []
            # 遍历每一列
            for col in df.columns:
                for cell in df[col].dropna():
                    cell = str(cell).zfill(5)

                    if isinstance(cell, str):
                        # 使用正则表达式提取数字
                        codes = re.findall(r'\b\d+\b', cell)
                        for code in codes:
                            if code == '00000':
                                continue
                            data.append(code)
            
                code_property_map[sheet_name] = data
            
        def get_property(code):
            if len(str(code)) != 5:
                return '邮编错误,不足五位'
            for property_name, codes in code_property_map.items():
                if str(code) in codes:
                    return property_name
            return 'Unknown'
            
        input_df['property'] = input_df[first_column_name].apply(get_property)
        return input_df
    

def get_ups_zip_data():    
# 获取property定义Excel文件路径
    property_excel_path = os.path.join(os.getcwd(), 'file', 'remoteaddresscheck', 'area-surcharge-zips-us-en.xlsx')
    
    # 读取property定义Excel中的所有sheet
    xl = pd.ExcelFile(property_excel_path)
    code_property_map = {}
    
    # 遍历每个sheet获取code和property的对应关系
    for sheet_name in xl.sheet_names:
        df = pd.read_excel(property_excel_path, sheet_name=sheet_name)
        #将数字字符串转换为数字
        data = []
        # 遍历每一列
        for col in df.columns:
            for cell in df[col].dropna():
                cell = str(cell).zfill(5)
                if isinstance(str(cell), str):
                    # 使用正则表达式提取数字
                    codes = re.findall(r'\b\d+\b', cell)
                    for code in codes:
                        if code == '00000':
                            continue
                        data.append(code)
        
            code_property_map[sheet_name] = data

    return code_property_map

def query_usps_zip(zipcode):
    # USPS 邮编查询接口
    url = 'https://ziplucas.com/zipCodeClassify?zipcode='+zipcode
    

    
    # 构建请求头
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        # 发送POST请求
        response = requests.get(url, headers=headers)
        
        # 检查响应状态
        if response.status_code == 200:
            # 返回JSON数据
            result_json = response.json()
            result = {
                "defaultCity":result_json.get("city").upper(),
                "defaultState":result_json.get("state").upper(),
                "resultStatus":"SUCCESS"
            }
            return result
        else:
            print(f'请求失败,状态码: {response.status_code}')
            return None
            
    except Exception as e:
        print(f'发生错误: {traceback.format_exc()}')
        return None
    




class MinioClient:
    def __init__(self, endpoint, access_key, secret_key, bucket_name, secure=False):
        """
        初始化 MinIO 客户端。

        Args:
            endpoint (str): MinIO 服务器的地址，例如 'localhost:9000'。
            access_key (str): 访问密钥。
            secret_key (str): 秘密密钥。
            bucket_name (str): 存储桶名称。
            secure (bool): 是否使用 SSL。默认为 False。
        """
        self.endpoint = endpoint
        self.access_key = access_key
        self.secret_key = secret_key
        self.bucket_name = bucket_name
        self.secure = secure
        self.client = None

    def connect(self):
        """
        连接到 MinIO 服务器。
        """
        try:
            self.client = Minio(
                self.endpoint,
                access_key=self.access_key,
                secret_key=self.secret_key,
                secure=self.secure
            )
            logger.info(f"成功连接到 MinIO 服务器: {self.endpoint}")
            # 检查存储桶是否存在，如果不存在则创建
            if not self.client.bucket_exists(self.bucket_name):
                self.client.make_bucket(self.bucket_name)
                logger.info(f"存储桶 '{self.bucket_name}' 创建成功。")
        except S3Error as e:
            logger.error(f"连接到 MinIO 失败: {e}")
            self.client = None

    def upload_file(self, file_path, object_name=None):
        """
        上传文件到 MinIO 存储桶。

        Args:
            file_path (str): 要上传的文件的本地路径。
            object_name (str, optional): MinIO 中存储的对象名称。如果为 None，则使用文件名。
        """
        if not self.client:
            logger.error("MinIO 客户端未连接。请先调用 connect()。")
            return

        if object_name is None:
            object_name = os.path.basename(file_path)

        try:
            self.client.fput_object(
                self.bucket_name, object_name, file_path
            )
            logger.info(f"文件 '{file_path}' 成功上传到 '{self.bucket_name}/{object_name}'")
            return f"{self.bucket_name}/{object_name}"  # 返回对象名
        except S3Error as e:
            logger.error(f"上传文件 '{file_path}' 失败: {e}")
            return None

    def download_file(self, object_name, file_path=None):
        """
        从 MinIO 存储桶下载文件。

        Args:
            object_name (str): MinIO 中存储的对象名称。
            file_path (str, optional): 下载文件的本地路径。如果为 None，则返回 bytes。
        """
        if not self.client:
            logger.error("MinIO 客户端未连接。请先调用 connect()。")
            return None

        try:
            if file_path:
                self.client.fget_object(self.bucket_name, object_name, file_path)
                logger.info(f"文件 '{object_name}' 成功下载到 '{file_path}'")
                return file_path  # 返回本地文件路径
            else:
                response = self.client.get_object(self.bucket_name, object_name)
                file_bytes = response.data
                logger.info(f"文件 '{object_name}' 成功下载到 bytes")
                return file_bytes  # 返回文件 bytes
        except S3Error as e:
            logger.error(f"下载文件 '{object_name}' 失败: {e}")
            return None


