from datetime import datetime
import glob
import os
import re
import shutil
import zipfile
import openpyxl
import pandas as pd
from openpyxl import load_workbook, Workbook

class ExcelProcessor:
    def __init__(self):
        self.template_file = '时颖买单资料模块.xlsm'
        self.output_file = '时颖买单资料模块_processed.xlsx'
        self.wb = load_workbook(self.template_file)
        self.fapiao_df = None


    def clean_ninth_column_data(self, df):
        """清理第9列数据"""
        # 假设第九列列名为 'I'（0-based索引对应列名）
        df['I'] = df['I'].apply(lambda x: x if pd.isna(x) else '|'.join(
            [part for part in str(x).split('|') 
             if part != "无品牌" and re.search(r'[A-Za-z]', part)]
        ))
        return df

    def process_data(self, df:pd.DataFrame):
        """数据聚合和计算"""
        #清除品名位空的行
        df = df[df['品名'].notna()]
        df.loc[:, '品牌'] = df['品牌'].fillna("")
        df.loc[:, '要素备注'] = df['要素备注'].apply(lambda x:x.split("无品牌")[0] if "无品牌" in x else x)
        # 分组字段
        group_cols = ['织造方式', '品名', '要素备注', '英文', '品牌', '海关编码']
        # 聚合字段
        agg_funcs = {
            '数量': 'sum',
            '重量': 'sum',
            '规格': 'count'
        }
        # 分组聚合
        grouped = df.groupby(group_cols).agg(agg_funcs).reset_index()
        
        # 新增计算列
        grouped['unit'] = grouped['品名'].apply(
            lambda x: 'suit' if str(x).endswith("套装") else 'pcs'
        )
        grouped['unit_price'] = grouped['品名'].apply(lambda x: self.fapiao_df[self.fapiao_df['品名'] == x]['单价(USD)'].iloc[0] if not self.fapiao_df[self.fapiao_df['品名'] == x]['单价(USD)'].empty else 0)
        grouped['total_price'] = grouped['数量'] * grouped['unit_price']
        grouped['dimension'] = 'ctn'
        grouped['net_weight'] = grouped['重量'] * 0.9
        grouped['carton_volume'] = grouped['规格'] * 0.08
        
        # 选择输出列并重命名
        output_cols = {
            '海关编码': 'customs_code',
            '品名': 'product_name',
            '英文': 'english_name',
            '数量': 'quantity',
            'unit': 'unit',
            'unit_price': 'unit_price',
            'total_price': 'total_price',
            '规格': 'carton',
            'dimension': 'dimension',
            '重量': 'gross_weight',
            'net_weight': 'net_weight',
            'carton_volume': 'cubic_volume',
            '要素备注': 'declaration_elements',
            '品牌': 'brand'
        }
        result = grouped.rename(columns=output_cols)[list(output_cols.values())]
        return result



    def generate_summary_table(self, data:dict):
        """生成生产生成器总表"""
        # 读取数据区工作表
        output_sheet = self.wb['输出区']
        
        # 从第20行开始写入数据
        row_index = 20
        start_row = row_index
        for row_data in data:
            # 将数据写入Excel表格
            output_sheet.cell(row=row_index, column=1).value = row_data.get('customs_code', '')
            output_sheet.cell(row=row_index, column=2).value = row_data.get('product_name', '')
            output_sheet.cell(row=row_index, column=4).value = row_data.get('english_name', '')
            output_sheet.cell(row=row_index, column=5).value = row_data.get('quantity', '')
            output_sheet.cell(row=row_index, column=6).value = row_data.get('unit', '')
            output_sheet.cell(row=row_index, column=7).value = row_data.get('unit_price', '')
            output_sheet.cell(row=row_index, column=8).value = row_data.get('total_price', '')
            output_sheet.cell(row=row_index, column=10).value = row_data.get('carton', '')
            output_sheet.cell(row=row_index, column=11).value = row_data.get('dimension', '')
            output_sheet.cell(row=row_index, column=12).value = row_data.get('gross_weight', '')
            output_sheet.cell(row=row_index, column=14).value = row_data.get('net_weight', '')

            output_sheet.cell(row=row_index, column=15).value = row_data.get('cubic_volume', '')
            output_sheet.cell(row=row_index, column=16).value = row_data.get('declaration_elements', '')
            output_sheet.cell(row=row_index, column=17).value = row_data.get('brand', '')
            
            row_index += 1
            
        # 合并合计行的单元格
        end_row = row_index - 1
        # output_sheet.merge_cells(f'J{20}:J{end_row}')
        # output_sheet.merge_cells(f'L{20}:L{end_row}')
        # output_sheet.merge_cells(f'N{20}:N{end_row}')
        
        # 使用Excel公式计算合计值
        output_sheet.cell(row=row_index, column=10).value = f'=SUM(J{start_row}:J{end_row})'  # 箱数合计
        output_sheet.cell(row=row_index, column=12).value = f'=SUM(L{start_row}:L{end_row})'  # 毛重合计
        output_sheet.cell(row=row_index, column=14).value = f'=SUM(N{start_row}:N{end_row})'  # 净重合计
        output_sheet.cell(row=row_index, column=15).value = f'=SUM(O{start_row}:O{end_row})'  # 体积合计


    def import_packing_list(self, file_path):
        """导入装箱单"""
       
        # 读取源文件
        columns = ['品名', '英文', '织造方式', '海关编码', '成分', '填充物', '客户Sku', '数量', '要素备注', '重量', '规格', 'ASIN', '品牌']
        src_df = pd.read_excel(file_path).dropna(axis=1, how='all')
        src_df = src_df.iloc[:, 1:]
        
        # 检查并删除不在columns[2:]中的列
        existing_cols = src_df.columns.tolist()
        for col in existing_cols[2:]:
            if col not in columns[2:]:
                src_df = src_df.drop(columns=col)
                
        # 如果没有品牌列则添加
        if '品牌' not in src_df.columns:
            src_df['品牌'] = ''

        src_df.columns = columns
        
        # 删除空行
        src_df = src_df.dropna(subset=[src_df.columns[0]])

        processed_df = self.process_data(src_df)
        dict_data = processed_df.to_dict(orient='records')

        return dict_data
    

    def brand_process(self, brand_file_path,xiangdian_file_path):
        """品牌处理"""
        # 读取源文件
        headers = ['FBA订单号','amazon reference','number','品牌']
        brand_df = pd.read_excel(brand_file_path, skiprows=1, names=headers)
        wb = openpyxl.load_workbook(xiangdian_file_path)
        # 读取箱单文件
        sheet = wb.active
        
        # 检查是否存在品牌列,如果不存在则添加
        has_brand_column = False
        for cell in sheet[1]:
            if cell.value == '品牌':
                has_brand_column = True
                break
                
        if not has_brand_column:
            # 获取最后一列的列号
            last_col = sheet.max_column
            # 在最后一列添加品牌列
            sheet.cell(row=1, column=last_col + 1).value = '品牌'
            brand_col_index = last_col
        else:
            # 找到品牌列的索引
            for idx, cell in enumerate(sheet[1]):
                if cell.value == '品牌':
                    brand_col_index = idx
                    break
        
        current_fba = None
        current_brand = None
        
        # 遍历所有行
        for row in sheet.iter_rows(min_row=1):
            cell_value = row[0].value
            
            # 如果是FBA开头的新订单号
            if isinstance(cell_value, str) and cell_value.startswith('FBA'):
                current_fba = cell_value
                # 在brand_df中查找对应的品牌
                brand_row = brand_df[brand_df['FBA订单号'] == current_fba]
                if not brand_row.empty:
                    current_brand = brand_row.iloc[0]['品牌']
                else:
                    current_brand = None
                    
            # 如果有当前品牌,填入品牌列
            if current_brand and not str(cell_value).startswith('FBA'):
                row[brand_col_index].value = current_brand
                
        # 保存修改后的文件
        wb.save(xiangdian_file_path)
    def qinggguan_fapiao_process(self, fapiao_file_path):
        """清关发票"""
        # 读取源文件
        fapiao_df = pd.read_excel(fapiao_file_path,sheet_name="发票",skiprows=19).dropna(axis=1, how='all')
        shuju_df = pd.read_excel(fapiao_file_path,sheet_name="数据",skiprows=6,usecols="A,D").dropna(axis=0, how='all')
        shuju_df = shuju_df.dropna(subset=[shuju_df.columns[0]])
        
        # 重命名列以便合并
        shuju_df.columns = ['品名', '货物描述']
        
        # 删除Total Quantity及以后的行
        total_quantity_idx = fapiao_df[fapiao_df.iloc[:,0] == 'Total Quantity'].index 
        total_quantity = fapiao_df.iloc[total_quantity_idx[0]]['Unnamed: 2']
        totoal_package = fapiao_df.iloc[total_quantity_idx[0]]['单位']
        total_value = fapiao_df.iloc[total_quantity_idx[0]]['价值(USD)']
        if len(total_quantity_idx) > 0:
            fapiao_df = fapiao_df.iloc[:total_quantity_idx[0]]
            
        # 检查并打印列名,以便调试

        
        # 根据品名合并两个DataFrame
        fapiao_df = pd.merge(fapiao_df, shuju_df[['品名','货物描述']], 
                           left_on='货物描述', 
                           right_on='货物描述', 
                           how='left',
                           suffixes=('_发票', '_数据'))
        
        self.fapiao_df = fapiao_df
        return fapiao_df,total_quantity,totoal_package,total_value

    def generate_customs_doc(self):
        """生成报关文档"""
        # 读取模板文件并删除工作表
        if '数据区' in self.wb.sheetnames:
            del self.wb['数据区']
        if '操作区' in self.wb.sheetnames:
            del self.wb['操作区']
        
        # 读取托书工作表的carton_num
        carton_sheet = self.wb['托书']
        carton_num = carton_sheet.cell(row=18, column=3).value  # 第18行第3列
        
        # 保存到桌面
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        output_file = os.path.join(desktop_path, f"时颖{carton_num}ctns.xlsx")
        self.wb.save(output_file)
        print(f"报关文档已生成，保存到桌面：{output_file}")

def main(zip_path):
    try:
        
        # 获取zip文件名(不含扩展名)作为临时目录名
        temp_dir = os.path.splitext(os.path.basename(zip_path))[0]
        temp_path = os.path.join("temp", temp_dir + "_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        os.makedirs(temp_path, exist_ok=True)
    
        # 解压到临时目录
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            for file in zip_ref.namelist():
                try:
                    # 尝试使用cp437解码,然后用gbk编码
                    filename = file.encode('cp437').decode('gbk')
                except:
                    # 如果失败则保持原样
                    filename = file
                zip_ref.extract(file, temp_path)
                # 重命名为正确的中文名
                if file != filename:
                    os.rename(os.path.join(temp_path, file), 
                            os.path.join(temp_path, filename))
            
        # 查找装箱清单excel文件
        packing_files = glob.glob(os.path.join(temp_path, "*装箱清单*.xls*"))
        if not packing_files:
            raise Exception("未找到装箱清单文件")
        packing_file = packing_files[0]
        
        # 查找编码excel文件
        code_files = glob.glob(os.path.join(temp_path, "*编码*.xls*"))
        if not code_files:
            raise Exception("未找到编码文件") 
        fapiao_file = glob.glob(os.path.join(temp_path, "*清关发票*.xls*"))
        if not fapiao_file:
            raise Exception("未找到清关发票文件") 
        fapiao_file = fapiao_file[0]
        code_file = code_files[0]
        processor = ExcelProcessor()
        fapiao_df,total_quantity,totoal_package,total_value = processor.qinggguan_fapiao_process(fapiao_file_path=fapiao_file)

        processor.brand_process(brand_file_path=code_file,xiangdian_file_path=packing_file)

        # 导入装箱单
        import_data = processor.import_packing_list(file_path=packing_file)
        
        # 处理编码文件

        # 生成汇总表
        processor.generate_summary_table(import_data)
        # 生成报关文档
        processor.generate_customs_doc()
    finally:
        # 删除临时目录
        shutil.rmtree(temp_path)

if __name__ == "__main__":
    main(zip_path = r"C:\Users\a1337\Downloads\ABQ2 红色.zip")