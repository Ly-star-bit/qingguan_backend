from datetime import date, datetime, timedelta
from io import BytesIO
import os
from pathlib import Path
import random
import subprocess
import sys
import textwrap
import traceback
from uuid import UUID, uuid4
import bcrypt
from fastapi import Depends, FastAPI, HTTPException, Query,Request
from fastapi.responses import FileResponse,JSONResponse, StreamingResponse
import httpx
import jwt
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

from pydantic import BaseModel, validator
from typing import List, Optional, Union
import requests
from sqlalchemy import Boolean, Column, DateTime, Enum, Float, ForeignKey, String, Text, func
from sqlmodel import Relationship, SQLModel, Field, create_engine, Session, select
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.exc import OperationalError
from tenacity import retry, stop_after_attempt, wait_fixed
from sqlalchemy.pool import QueuePool
from loguru import logger
from dotenv import load_dotenv
from sqlalchemy.exc import InvalidRequestError
import comtypes.client
import casbin
from casbin_sqlalchemy_adapter import Adapter,CasbinRule
from starlette.middleware.base import BaseHTTPMiddleware
import jpype
import os
import time
from sqlalchemy.dialects.postgresql import UUID as SQLAlchemyUUID
from minio import Minio
from minio.error import S3Error
from openpyxl import Workbook as Openpyxl_Workbook
load_dotenv()


# 启动 JVM 并确保 JVM 启动在导入之前
if not jpype.isJVMStarted():
        jpype.startJVM()

# 在 JVM 启动后导入 Java 依赖的模块
from asposecells.api import Workbook, License, PdfSaveOptions,TextAlignmentType, SaveFormat,SheetSet


def upload_pdf_to_minio(pdf_file_path, object_name):
    """
    上传 PDF 文件到 MinIO 服务器。

    :param pdf_file_path: 本地 PDF 文件路径
    :param object_name: 上传到 MinIO 后的对象名称
    :return: 上传成功返回 True，否则返回 False
    """
    # 加载 .env 文件中的环境变量
    load_dotenv()

    # 从环境变量中获取 MinIO 配置信息
    server_url = os.getenv("MINIO_SERVER_URL")
    access_key = os.getenv("MINIO_ACCESS_KEY")
    secret_key = os.getenv("MINIO_SECRET_KEY")
    bucket_name = os.getenv("MINIO_BUCKET_NAME")

    try:
        # 初始化 MinIO 客户端
        client = Minio(
            server_url,
            access_key=access_key,
            secret_key=secret_key,
            secure=server_url.startswith("https")
        )

        # 确保桶存在
        if not client.bucket_exists(bucket_name):
            client.make_bucket(bucket_name)

        # 上传文件
        client.fput_object(
            bucket_name, object_name, pdf_file_path,
        )
        print(f"'{pdf_file_path}' is successfully uploaded as '{object_name}' to bucket '{bucket_name}'.")
        return True
    except S3Error as err:
        print(f"An error occurred: {err}")
        return False
def excel2pdf(excel_path: str, pdf_save_path: str = None) -> str:
    # 加载License文件
    apcelllic = License()
    apcelllic.setLicense('JAVA-Aspose.Excel-24.7/license.xml')

    # 打开Excel文件
    wb = Workbook(excel_path)
    # 删除名称为 "Evaluation Warning" 的工作表（如果存在）
    sheets = wb.getWorksheets()
    eval_warning_sheet = sheets.get("Evaluation Warning")
    if eval_warning_sheet is not None:
        sheets.removeAt("Evaluation Warning")
    # 配置PDF保存选项
    saveOption = PdfSaveOptions()
    
    # 确保每个工作表单独保存为一个PDF页面
    # saveOption.setOnePagePerSheet(True)  # 如果为True，将整个工作表压缩到一个PDF页面上
    saveOption.setAllColumnsInOnePagePerSheet(True) #所有列在一页，但是可能行在多页

    # 计算公式
    saveOption.setCalculateFormula(True)  # 计算公式并将其值保存在PDF中

    # 设置字体相关选项
    saveOption.setCheckWorkbookDefaultFont(True)  # 检查工作簿的默认字体，以避免出现方块字符
    saveOption.setCheckFontCompatibility(True)  # 检查每个字符的字体兼容性
    saveOption.setDefaultFont("Arial")  # 设置默认字体（如果未设置正确的字体）

    # 设置图像处理
    saveOption.setImageResample(220, 85)  # 设置图像的PPI和JPEG质量，减少PDF文件大小

    # 设置其他相关选项
    saveOption.setEmbedStandardWindowsFonts(True)  # 嵌入标准的Windows字体
    saveOption.setClearData(False)  # 在保存后不清除工作簿的数据
    saveOption.setCompliance(0)  # 设置PDF标准合规级别，如需要合规的PDF/A等格式
    saveOption.setDisplayDocTitle(True)  # 在PDF窗口的标题栏显示文档标题

    # 如果没有指定保存路径，则使用与 Excel 文件相同的路径
    if pdf_save_path is None:
        pdf_save_path = os.path.dirname(excel_path)
    
    # 获取Excel文件的文件名（不含扩展名）
    excel_name = os.path.splitext(os.path.basename(excel_path))[0]
    
    # 设置PDF文件的完整保存路径
    pdf_file = os.path.join(pdf_save_path, f"{excel_name}.pdf")

    # 保存为PDF
    wb.save(pdf_file, saveOption)
    
    return pdf_file


# 配置日志文件保存
logger.add("app.log",
           format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}",
           level="INFO",
           rotation="1 MB",
           retention="10 days",
           compression="zip")

# 配置 loguru 输出到控制台
# 定义你的数据库连接配置
DATABASE_CONFIG = {
    'user': os.getenv("MYSQL_USER"),
    'password': os.getenv("MYSQL_PASS"),
    'host': os.getenv("MYSQL_HOST"),
    'database': os.getenv("MYSQL_DB"),
    "port": int(os.getenv("MYSQL_PORT"))
}

# 创建数据库连接字符串
DATABASE_URL = f"mysql+mysqlconnector://{DATABASE_CONFIG['user']}:{DATABASE_CONFIG['password']}@{DATABASE_CONFIG['host']}:{DATABASE_CONFIG['port']}/{DATABASE_CONFIG['database']}"
engine = create_engine(f'mysql+pymysql://{DATABASE_CONFIG["user"]}:{DATABASE_CONFIG["password"]}@{DATABASE_CONFIG["host"]}:{DATABASE_CONFIG["port"]}/{DATABASE_CONFIG["database"]}')

adapter = Adapter(engine)
enforcer = casbin.Enforcer('model.conf', adapter)
engine = create_engine(
    DATABASE_URL,
    pool_size=10,           # 连接池大小
    max_overflow=20,        # 连接池溢出大小
    pool_timeout=30,        # 连接池超时时间
    pool_recycle=1800,      # 连接池回收时间
    poolclass=QueuePool,     # 使用QueuePool连接池
    pool_pre_ping=True,  # 新增

)
# 密钥和算法
SECRET_KEY = "3c8a8a0b4e4b3e5e8a9d10b5f4b6a9c7e2a1f0b6a3c4e9d8b7c6d5a1e2f3a4b5"
ALGORITHM = "HS256"
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def create_refresh_token(data: dict):
    expire = datetime.utcnow() + timedelta(days=7)
    to_encode = data.copy()
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt
@retry(stop=stop_after_attempt(3), wait=wait_fixed(2), reraise=True)
def get_session():
    try:
        return Session(engine)
    except OperationalError as e:
        logger.error(f"Database connection error: {e}")
        raise
TypeEnum = Enum('空运', '海运', '混合', name='type_enum')

class Port(SQLModel, table=True):
    __tablename__ = 'port'
    id: int = Field(default=None, primary_key=True)
    port_name: str = Field(max_length=255, nullable=False)
    sender_name: str = Field(max_length=255, default="")
    receiver_name: str = Field(max_length=255, nullable=False)
class Product3(SQLModel, table=True):
    总税率: str = Field(sa_column=Column("总税率", String(255), ), default="")
    中文品名: str = Field(sa_column=Column("中文品名", String(255), ), default="")
    英文品名: str = Field(sa_column=Column("英文品名", String(255), ), default="")
    HS_CODE: str = Field(sa_column=Column("HS_CODE", String(255), ), default="")
    Duty: str = Field(sa_column=Column("Duty(%)", String(255), ), default="")
    加征: str = Field(sa_column=Column("加征%", String(255), ), default="")
    一箱税金: str = Field(sa_column=Column("一箱税金", String(255), ), default="")
    豁免代码: str = Field(sa_column=Column("豁免代码", String(255), ), default="")
    豁免代码含义: str = Field(sa_column=Column("豁免代码含义", Text, ), default="")

    豁免截止日期说明: str = Field(sa_column=Column("豁免截止日期/说明", String(255), ), default="")
    豁免过期后: str = Field(sa_column=Column("豁免过期后", String(255), ), default="")
    认证: str = Field(sa_column=Column("认证？", String(255), ), default="")
    件箱: str = Field(sa_column=Column("件/箱", String(255), ), default="")
    单价: str = Field(sa_column=Column("单价", String(255), ), default="")
    材质: str = Field(sa_column=Column("材质", String(255), ), default="")
    用途: str = Field(sa_column=Column("用途", String(255), ), default="")
    更新时间: datetime = Field(sa_column=Column("更新时间", DateTime), default_factory=lambda: datetime.utcnow())
    类别: str = Field(sa_column=Column("类别", String(255), ), default="")
    属性绑定工厂: str = Field(sa_column=Column("属性绑定工厂", String(255), ), default="")
    序号: Optional[int] = Field(default=None, primary_key=True)
    备注: str = Field(sa_column=Column("备注", String(255), ), default="")
    单件重量合理范围: str = Field(sa_column=Column("单件重量合理范围", String(255), ), default="")
    客户: str = Field(sa_column=Column("客户", String(255), ), default="")
    报关代码: str = Field(sa_column=Column("报关代码", String(255), ), default="")
    客人资料美金: str = Field(sa_column=Column("客人资料美金", String(255), ), default="")

    single_weight: float = Field(sa_column=Column("single_weight", Float(), ))
    自税:bool = Field(sa_column=Column("自税", Boolean(), ))
    类型: str = Field(sa_column=Column("类型", TypeEnum, default="混合"))

    @validator('更新时间', pre=True, always=True)
    def parse_datetime(cls, value):
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value.replace('Z', '+00:00'))
            except ValueError:
                return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        return value

    class Config:
        json_encoders = {
            datetime: lambda v: v.strftime('%Y-%m-%d %H:%M:%S')
        }
class HaiYunZiShui(SQLModel, table=True):
    __tablename__ = '海运自税'
    id: int = Field(default=None, primary_key=True, description="自增主键")
    zishui_name: str = Field(default=None, max_length=100, description="自税名称")
    sender: str = Field(default=None, description="发货人")
    receiver: str = Field(default=None, description="收货人")
# 定义shippersandreceivers表的数据模型，添加主键字段
class ShippersAndReceivers(SQLModel, table=True):
    __tablename__ = 'shippersandreceivers'
    id: Optional[int] = Field(default=None, primary_key=True)
    ShipperName: Optional[str] = None
    ShipperAddress: Optional[str] = None
    ReceiverName: Optional[str] = None
    ReceiverAddress: Optional[str] = None
    Attribute: Optional[str] = None
    ChineseName: Optional[str] = None
    EnglishName: Optional[str] = None
    Address: Optional[str] = None

    __table_args__ = {'comment': 'Shippers and Receivers table'}

class FactoryData(SQLModel, table=True):
    __tablename__ = "工厂数据"
    id: Optional[int] = Field(default=None, primary_key=True)
    属性: str = Field(max_length=255, nullable=False)
    中文名字: str = Field(max_length=255, nullable=False)
    英文: str = Field(nullable=False)
    地址: str = Field(nullable=False)
class ConsigneeData(SQLModel, table=True):
    __tablename__ = "收发货人"
    id: Optional[int] = Field(default=None, primary_key=True)
    中文: str = Field(max_length=255, nullable=False)
    发货人: str = Field(max_length=255, nullable=False)
    发货人详细地址: str = Field(nullable=False)
    类型: str  = Field(nullable=False)
    备注: str = Field(nullable=False)
    hide: str = Field(nullable=False)



class IpWhiteList(SQLModel, table=True):
    __tablename__ = "ip_white_list"
    id: Optional[int] = Field(default=None, primary_key=True)
    ip: str = Field(max_length=255, nullable=False)
    remarks: str = Field(max_length=255, nullable=False)




class Dalei(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    hs_code: Optional[str] = Field(default=None, max_length=255)
    英文大类: Optional[str] = Field(default=None, max_length=255)
    中文大类: Optional[str] = Field(default=None, max_length=255)

class Policy(BaseModel):
    sub: str
    obj: str
    act: str
    eft: str = None
class FileInfo(BaseModel):
    name: str
    time: datetime
class UpdatePolicy(BaseModel):
    old_sub: str
    old_obj: str
    old_act: str
    old_eft: str
    new_sub: str
    new_obj: str
    new_act: str
    new_eft: str
class update_cumstom_clear_history_summary_remarks(BaseModel):
    remarks:str
    id:str
class Group(BaseModel):
    user: str
    group: str
class User(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    username: str = Field(index=True, unique=True, nullable=False)
    password: str = Field(nullable=False)
    created_at: datetime = Field(default_factory=datetime.utcnow)

class UserCreate(BaseModel):
    username: str
    password: str
    permissions: List[str]
class UserLogin(BaseModel):
    username: str
    password: str
class UserUpdate(BaseModel):
    username: Optional[str]
    password: Optional[str]
    permissions: Optional[List[str]]
    # 定义白名单
class IPWhitelistMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        client_ip = request.client.host
        print(client_ip)
        # 从数据库中获取白名单
        try:
            session = get_session()
            with session:
                statement = select(IpWhiteList.ip)
                result = session.exec(statement)
                ip_whitelist = [row for row in result] + ['127.0.0.1','localhost']
        except Exception as e:
            logger.error(f"Error fetching IP whitelist from database: {e}")
            ip_whitelist = []

        if client_ip not in ip_whitelist:
            return JSONResponse(
                status_code=403,
                content={"detail": f"IP {client_ip} not allowed"}
            )
        
        response = await call_next(request)
        return response
# 定义FastAPI应用
app = FastAPI()
# 设置CORS
origins = [
    "http://192.168.20.87:3000",
    "http://192.168.20.143:8088",
    "http://localhost:3000",
    "http://47.103.138.130:3000",
    "http://47.103.138.130:3001"

    # 你可以在这里添加其他允许的源
]
app.add_middleware(IPWhitelistMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def start_nextjs():
    # 获取当前工作目录
    cwd = os.getcwd()
    # 进入到 Next.js 项目的目录
    logger.info("正在启动服务next")
   
    if not os.path.isdir(cwd):
        raise ValueError(f"Next.js directory does not exist: {cwd}")
    # 启动 Next.js 应用
    process = subprocess.Popen(['next', 'start', '-H', '0.0.0.0'], cwd=cwd, shell=True)
    return process
def create_db_and_tables():
    SQLModel.metadata.create_all(engine)
# # # 在启动 FastAPI 应用之前启动 Next.js
@app.on_event("startup")
async def startup_event():
    create_db_and_tables()
    # start_nextjs()
# 定义请求体数据模型
@app.on_event("shutdown")
def shutdown_event():
    # 关闭 JVM
    if jpype.isJVMStarted():
        jpype.shutdownJVM()
# 定义请求体数据模型
class ProductData(BaseModel):
    product_name: str
    box_num: int
    single_price: float = None
    packing: Union[int, float] = None



class ShippingRequest(BaseModel):
    predict_tax_price:float=0.00
    totalyugutax: float=0.00
    port:str
    packing_type:str
    shipper_name: str
    receiver_name: str
    master_bill_no: str
    gross_weight: float
    volume: float
    product_list: List[ProductData]
class DaleiCreate(BaseModel):
    hs_code: Optional[str]
    英文大类: Optional[str]
    中文大类: Optional[str]
# 定义处理数据的函数
def process_shipping_data(
    shipper_name: str, 
    receiver_name: str, 
    master_bill_no: str, 
    gross_weight: int, 
    volume: int, 
    product_list: List[ProductData],
    totalyugutax:float,
    predict_tax_price:float
    ):
    # 创建引擎
    with get_session() as session:
        num_products = sum([i.box_num for i in product_list])
        if num_products == 0:
            raise ValueError("Product list cannot be empty")



        avg_gross_weight = gross_weight / num_products
        avg_volume = volume / num_products

        results = []
        accumulated_gross_weight = 0
        accumulated_volume = 0
        if product_list[0].single_price or product_list[0].packing:

            execute_type = "Sea"


        else:

            execute_type = "Air"

        # 查询发货人地址
        shipper_query = select(ConsigneeData).where(ConsigneeData.发货人 == shipper_name)
        shipper_record = session.exec(shipper_query).first()

        # 查询收件人地址
        receiver_query = select(ConsigneeData).where(ConsigneeData.发货人== receiver_name)
        receiver_record = session.exec(receiver_query).first()

        if not shipper_record:
            raise ValueError(f"Shipper '{shipper_name}' not found in database")
        if not receiver_record:
            raise ValueError(f"Receiver '{receiver_name}' not found in database")

        shipper_address = shipper_record.发货人详细地址
        receiver_address = receiver_record.发货人详细地址
        detail_data_log_list = []
        def safe_round(value, default=0.0):
            try:
                return round(float(value), 4)
            except (ValueError, TypeError):
                return default

        for idx, product in enumerate(product_list):
            product_name = product.product_name
            box_num = product.box_num

            
            product_query = select(Product3).where(Product3.中文品名 == product_name)
            product_record = session.exec(product_query).first()
            duty = safe_round(product_record.Duty)
            additional_duty = safe_round(product_record.加征)
            detail_data_log = CustomClearHistoryDetailLog(
                hs_code=product_record.HS_CODE,
                chinese_name=product_record.中文品名,
                transport_mode=execute_type,
                master_bill_number=master_bill_no,
                total_tax_rate=duty + additional_duty,
                exemption_code=product_record.豁免代码,
                category=product_record.类别,
            )
            detail_data_log_list.append(detail_data_log)
            if not product_record:
                raise ValueError(f"Product '{product_name}' not found in database")
            if product.single_price:
                single_price = product.single_price
            else:
                single_price = product_record.单价

            if product.packing:
                packing = product.packing
            else:
                packing = product_record.件箱
            product_type = session.exec(select(Product3.属性绑定工厂).where(Product3.HS_CODE == product_record.HS_CODE)).first()

            address_name_list = session.exec(
                select(FactoryData.地址, FactoryData.英文)
                .where(
                    (FactoryData.地址 != None) & 
                    (FactoryData.属性 == product_type)
                )
            ).all()

            if not address_name_list:
                logger.warning("地址数据库中没有对应的属性")
                return {"product_attribute": "不存在"}

            random_address_name = random.choice(address_name_list)
            address = random_address_name[0]
            address_name = random_address_name[1]

            if idx == len(product_list) - 1:  # 处理最后一个产品
                gross_weight_for_this_product = round(gross_weight - accumulated_gross_weight, 2)
                volume_for_this_product = round(volume - accumulated_volume, 2)
            else:

               
                gross_weight_for_this_product = round(avg_gross_weight * box_num, 2)
                volume_for_this_product = round(avg_volume * box_num, 2)
                accumulated_gross_weight += gross_weight_for_this_product
                accumulated_volume += volume_for_this_product
            # 如果已经有固定的单件KG，则net weight不需要计算
            if product_record.single_weight:
                net_weight_for_this_product = product_record.single_weight
                logger.info(f"{product_record.中文品名}->净重为{net_weight_for_this_product}，实重为{gross_weight_for_this_product}")
                #净重 不能大于 实际重量
                if net_weight_for_this_product > gross_weight_for_this_product:
                    logger.error(f"{product_record.中文品名}-> net weight 大于 gross weight")
                    return {"msg": f"{product_record.中文品名}-> net weight 大于 gross weight","type":"net_weight大于gross_weight"}
                
            else:
                net_weight_for_this_product = round(gross_weight_for_this_product * 0.8, 2)

            product_data = {
                'MasterBillNo': master_bill_no,
                "shipper_name": shipper_name,
                "shipper_address": shipper_address,
                "receiver_address": receiver_address,
                "receiver_name": receiver_name,
                'ProductName': product_name,
                'carton': box_num,
                # 'quanity': box_num * float(product_record.件箱),
                'quanity': box_num * float(packing),
                "danwei": "PCS" if product_record.HS_CODE else '',
                # "unit_price": product_record.单价,
                "unit_price": single_price,
                # 'total_price': float(product_record.单价) * box_num * float(product_record.件箱),
                'total_price': float(single_price) * box_num * float(packing),

                'HS_CODE': product_record.HS_CODE,
                'DESCRIPTION': product_record.英文品名,
                'GrossWeight': gross_weight_for_this_product,
                'net_weight': net_weight_for_this_product,
                'Volume': volume_for_this_product,
                "usage": product_record.用途,
                "texture": product_record.材质,
                "address_name": address_name or "",
                "address": address or "",
                "note": product_record.豁免代码,
                "note_explaination": product_record.豁免代码含义,
                'execute_type':execute_type

            }

            results.append(product_data)
        
        summary_log_data = CustomClearHistorySummaryLog(
            filename=master_bill_no,
            generation_time=datetime.now(),
            port="",
            packing_type="",
            shipper = shipper_name,
            consignee= receiver_name,
            estimated_tax_amount=totalyugutax,
            gross_weight_kg=gross_weight,
            volume_cbm= volume,
            total_boxes=sum([i.box_num for i in product_list]),
            estimated_tax_rate_cny_per_kg=predict_tax_price,
            details=detail_data_log_list

        )
        return results,summary_log_data

def generate_excel_from_template_test(data,totalyugutax):
    apcelllic = License()
    apcelllic.setLicense('JAVA-Aspose.Excel-24.7/license.xml')
    template_path = "副本清关发票箱单模板 - 0918更新.xlsx"
    wb = Workbook(template_path)
    civ_sheet = wb.getWorksheets().get("CIV")
    pl_sheet = wb.getWorksheets().get("PL")
    huomian_explaination_sheet = wb.getWorksheets().get("豁免说明")
    start_row = 13

    def set_cell_value(sheet, row, column, value):
        cell = sheet.getCells().get(row - 1, column - 1)  # Adjust for zero-based index
        
        if isinstance(value, str) and value.startswith('='):
            # 如果 value 是一个公式
            cell.setFormula(value)
        else:
            # 否则，设置为普通值
            cell.putValue(value)
        
        style = cell.getStyle()
        style.setTextWrapped(True)
        style.setHorizontalAlignment(TextAlignmentType.CENTER)
        style.setVerticalAlignment(TextAlignmentType.CENTER)
        cell.setStyle(style)


        # Auto-adjust row height
        # Adjust based on value length, e.g., number of lines
        # num_lines = value.count("\n") + 1
        # if num_lines > 1:
        #     sheet.getCells().setRowHeight(row - 1, num_lines * 15)

    # Fill CIV content
    set_cell_value(civ_sheet, 1, 1, data[0]["shipper_name"])
    set_cell_value(civ_sheet, 2, 1, data[0]["shipper_address"])
    set_cell_value(civ_sheet, 6, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

    today_minus_5 = datetime.now() - timedelta(days=5)
    formatted_date = today_minus_5.strftime("%Y%m%d")
    random_number = random.randint(1000, 9999)
    result_1 = f"{formatted_date}{random_number}"
    set_cell_value(civ_sheet, 6, 9, result_1)
    set_cell_value(civ_sheet, 7, 9, result_1)
    set_cell_value(civ_sheet, 8, 9, datetime.now().strftime("%Y/%m/%d"))

    set_cell_value(pl_sheet, 1, 1, data[0]["shipper_name"])
    set_cell_value(pl_sheet, 2, 1, data[0]["shipper_address"])
    set_cell_value(pl_sheet, 5, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

    set_cell_value(pl_sheet, 5, 9, result_1)
    set_cell_value(pl_sheet, 6, 9, result_1)
    set_cell_value(pl_sheet, 7, 9, datetime.now().strftime("%Y/%m/%d"))

    if data[0]["execute_type"] == "Sea":
        set_cell_value(civ_sheet, 9, 2, "")
        set_cell_value(civ_sheet, 10, 2, "US BY SEA")
        set_cell_value(pl_sheet, 8, 2, "")
        set_cell_value(pl_sheet, 9, 2, "US BY SEA")
    
    

    for index, item in enumerate(data):
        civ_row = 14 + index
        huomian_row = 5 + index
        pl_row = start_row + index


        # Fill CIV
        set_cell_value(civ_sheet, civ_row, 1, "=ROW()-ROW($A$14)+1")
        set_cell_value(civ_sheet, civ_row, 2, item["HS_CODE"])
        set_cell_value(civ_sheet, civ_row, 3, item["DESCRIPTION"])
        set_cell_value(civ_sheet, civ_row, 4, item["quanity"])
        set_cell_value(civ_sheet, civ_row, 5, item["danwei"])
        set_cell_value(civ_sheet, civ_row, 6, item["unit_price"])
        set_cell_value(civ_sheet, civ_row, 7, item["total_price"])
        set_cell_value(civ_sheet, civ_row, 8, item["texture"])
        set_cell_value(civ_sheet, civ_row, 9, item["address_name"])
        set_cell_value(civ_sheet, civ_row, 10, item["address"])
        set_cell_value(civ_sheet, civ_row, 11, item["note"])

        civ_sheet.autoFitRow(civ_row - 1)


        # Fill PL
        set_cell_value(pl_sheet, pl_row, 1, "=ROW()-ROW($A$13)+1")

        set_cell_value(pl_sheet, pl_row, 2, item["HS_CODE"])

        set_cell_value(pl_sheet, pl_row, 3, item["DESCRIPTION"])
        set_cell_value(pl_sheet, pl_row, 4, item["quanity"])
        set_cell_value(pl_sheet, pl_row, 5, item["danwei"])
        set_cell_value(pl_sheet, pl_row, 6, item["carton"])
        set_cell_value(pl_sheet, pl_row, 8, item["net_weight"])
        set_cell_value(pl_sheet, pl_row, 9, item["GrossWeight"])
        set_cell_value(pl_sheet, pl_row, 10, item["Volume"])
        pl_sheet.autoFitRow(pl_row - 1)

        # Fill 豁免说明
        set_cell_value(huomian_explaination_sheet, huomian_row, 1, item["HS_CODE"])
        set_cell_value(huomian_explaination_sheet, huomian_row, 2, item["DESCRIPTION"])
        set_cell_value(huomian_explaination_sheet, huomian_row, 3, item["usage"])
        set_cell_value(huomian_explaination_sheet, huomian_row, 4, item["note"])
        set_cell_value(huomian_explaination_sheet, huomian_row, 5, item["note_explaination"])
        huomian_explaination_sheet.autoFitRow(huomian_row - 1)

        if index  == len(data) - 1 :
            #如果是最后一个循环的数据，则不需要再添加一行了
            break
        # 在每个循环结束时增加一行，以避免覆盖
        # print(civ_sheet.getCells().get(civ_row, 2).getValue())
        if civ_sheet.getCells().get(civ_row + 1, 2).getValue() == "TOTAL":
            civ_sheet.getCells().insertRows(civ_row, 1)
        if pl_sheet.getCells().get(pl_row+ 1, 2).getValue() == "TOTAL":

            pl_sheet.getCells().insertRows(pl_row, 1)
        # pl_sheet.getCells().insertRows(pl_row, 1)

        # civ_sheet.getCells().insertRows(civ_row, 1)
        huomian_explaination_sheet.getCells().insertRows(huomian_row, 1)


    # Save the Excel file
    output_path = f"file/{time.time()}-{data[0]['MasterBillNo']} CI&PL-{totalyugutax}.xlsx"
    wb.calculateFormula()
    wb.save(output_path, SaveFormat.XLSX)
    print(f"Excel file generated: {output_path}")

        # 生成 PDF 文件
    pdf_path = excel2pdf(output_path, 'pdf')
    logger.info(f"pdf文件已成功生成: {pdf_path}")
    return pdf_path


def generate_excel_from_template(data):
    template_path = r"清关发票箱单模板.xlsx"
    # 读取模板文件
    wb = openpyxl.load_workbook(template_path)
    civ_sheet = wb["CIV"]
    pl_sheet = wb["PL"]
    huomian_explaination_sheet = wb['豁免说明']
    start_row = 13
    

    def set_cell_value(sheet, row, column, value):
        cell = sheet.cell(row=row, column=column)
        # 检查是否为合并单元格，如果是，只在左上角单元格写入值
        if cell.coordinate in sheet.merged_cells:
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left_cell = merged_range.start_cell
                    sheet[top_left_cell.coordinate].value = value
                    sheet[top_left_cell.coordinate].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    break
        else:
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # 自动调整行高
        wrap_length = 30  # 根据需要调整换行长度
        lines = textwrap.wrap(str(value), wrap_length)
        num_lines = len(lines)
        # 只有在内容需要换行时才调整行高
        if num_lines > 1:
            sheet.row_dimensions[row].height = num_lines * 15

 

    # 填充 CIV 一次内容
    set_cell_value(civ_sheet, 1, 1, data[0]["shipper_name"])
    set_cell_value(civ_sheet, 2, 1, data[0]["shipper_address"])
    set_cell_value(civ_sheet, 6, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

    today_minus_5 = datetime.now() - timedelta(days=5)
    formatted_date = today_minus_5.strftime("%Y%m%d")
    random_number = random.randint(1000, 9999)
    result_1 = f"{formatted_date}{random_number}"
    set_cell_value(civ_sheet, 6, 8, result_1)
    set_cell_value(civ_sheet, 7, 8, result_1)
    set_cell_value(civ_sheet, 8, 8, datetime.now().strftime("%Y/%m/%d"))

    set_cell_value(pl_sheet, 1, 1, data[0]["shipper_name"])
    set_cell_value(pl_sheet, 2, 1, data[0]["shipper_address"])
    set_cell_value(pl_sheet, 5, 1, f"{data[0]['receiver_name']}\n{data[0]['receiver_address']}")

    set_cell_value(pl_sheet, 5, 8, result_1)
    set_cell_value(pl_sheet, 6, 8, result_1)
    set_cell_value(pl_sheet, 7, 8, datetime.now().strftime("%Y/%m/%d"))

    if data[0]["execute_type"] == "Sea":
        set_cell_value(civ_sheet, 9, 2, "")
        set_cell_value(civ_sheet, 10, 2, "US BY SEA")
        set_cell_value(pl_sheet, 8, 2, "")
        set_cell_value(pl_sheet, 9, 2, "US BY SEA")

    for index, item in enumerate(data):
        civ_row = 14 + index
        huomian_row = 5 + index
        pl_row = start_row + index

        # 如果行数超出最大值，在最大行上方插入新行
        # civ_row = insert_row_if_needed(civ_sheet, civ_row, civ_max_row)
        # huomian_row = insert_row_if_needed(huomian_explaination_sheet, huomian_row, huomian_max_row)
        # pl_row = insert_row_if_needed(pl_sheet, pl_row, pl_max_row)

        # 填充 CIV
        set_cell_value(civ_sheet, civ_row, 1, item["HS_CODE"])
        set_cell_value(civ_sheet, civ_row, 2, item["DESCRIPTION"])
        set_cell_value(civ_sheet, civ_row, 3, item["quanity"])
        set_cell_value(civ_sheet, civ_row, 4, item["danwei"])
        set_cell_value(civ_sheet, civ_row, 5, item["unit_price"])
        set_cell_value(civ_sheet, civ_row, 6, item["total_price"])
        set_cell_value(civ_sheet, civ_row, 7, item["texture"])
        set_cell_value(civ_sheet, civ_row, 8, item["address_name"])
        set_cell_value(civ_sheet, civ_row, 9, item["address"])
        set_cell_value(civ_sheet, civ_row, 10, item["note"])

        # 填充 PL
        set_cell_value(pl_sheet, pl_row, 2, item["DESCRIPTION"])
        set_cell_value(pl_sheet, pl_row, 3, item["quanity"])
        set_cell_value(pl_sheet, pl_row, 4, item["danwei"])
        set_cell_value(pl_sheet, pl_row, 5, item["carton"])
        set_cell_value(pl_sheet, pl_row, 7, item["net_weight"])
        set_cell_value(pl_sheet, pl_row, 8, item["GrossWeight"])
        set_cell_value(pl_sheet, pl_row, 9, item["Volume"])

        # 填充豁免说明
        set_cell_value(huomian_explaination_sheet, huomian_row, 1, item["HS_CODE"])
        set_cell_value(huomian_explaination_sheet, huomian_row, 2, item["DESCRIPTION"])
        set_cell_value(huomian_explaination_sheet, huomian_row, 3, item["note"])
        set_cell_value(huomian_explaination_sheet, huomian_row, 4, item["note_explaination"])

    # 保存新的 Excel 文件
    output_path = f"file/{data[0]['MasterBillNo']} CI&PL.xlsx"
    wb.save(output_path)
    logger.info(f"excel文件已成功生成: {output_path}")

    # 生成 PDF 文件
    pdf_path = excel2pdf(output_path, 'pdf')
    logger.info(f"pdf文件已成功生成: {pdf_path}")
    
    return pdf_path




def excel_to_pdf(excel_path, pdf_path):
    # 创建 Excel 应用对象
    excel = comtypes.client.CreateObject('Excel.Application')
    # excel = comtypes.client.CreateObject('Ket.Application')  # 假设Ket是WPS的ProgID，具体请查阅WPS的文档

    excel.Visible = False

    # 打开 Excel 文件
    workbook = excel.Workbooks.Open(os.path.abspath(excel_path))

    try:
        # 循环遍历所有工作表
        for sheet in workbook.Sheets:
            # 计算有效数据范围
            max_row = sheet.UsedRange.Rows.Count
            max_col = sheet.UsedRange.Columns.Count

            # 设置页面大小
            sheet.PageSetup.PaperSize = 9  # 9 表示自定义页面大小
            sheet.PageSetup.LeftMargin = excel.InchesToPoints(0.5)  # 左边距设置为0.5英寸
            sheet.PageSetup.RightMargin = excel.InchesToPoints(0.5)  # 右边距设置为0.5英寸
            sheet.PageSetup.TopMargin = excel.InchesToPoints(0.5)  # 上边距设置为0.5英寸
            sheet.PageSetup.BottomMargin = 0  # 下边距设置为0

            # 确保页眉和页脚为空
            sheet.PageSetup.CenterHeader = ""
            sheet.PageSetup.CenterFooter = ""
            sheet.PageSetup.LeftHeader = ""
            sheet.PageSetup.LeftFooter = ""
            sheet.PageSetup.RightHeader = ""
            sheet.PageSetup.RightFooter = ""

            # 设置页面为横向
            sheet.PageSetup.Orientation = 2  # 2表示横向，1表示纵向
            sheet.PageSetup.Zoom = False
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = 1

        # 导出为 PDF，包含所有工作表
        workbook.ExportAsFixedFormat(0, os.path.abspath(pdf_path), IgnorePrintAreas=False)
        print(f"PDF file created at {pdf_path}")
    except Exception as e:
        print(f"Failed to export PDF: {e}")
    finally:
        workbook.Close(False)
        excel.Quit()



def shenzhen_customes_pdf_gennerate(data):
    apcelllic = License()
    apcelllic.setLicense('JAVA-Aspose.Excel-24.7/license.xml')
    template_path = "HAWB模板-空+海_测试新版.xls"
    wb = Workbook(template_path)
    shenzhn_sheet = wb.getWorksheets().get("S#-SZ-customs")

    def set_cell_value(sheet, row, column, value):
        cell = sheet.getCells().get(row - 1, column - 1)  # Adjust for zero-based index
        
        if isinstance(value, str) and value.startswith('='):
            # 如果 value 是一个公式
            cell.setFormula(value)
        else:
            # 否则，设置为普通值
            cell.putValue(value)
        
        style = cell.getStyle()
        style.setTextWrapped(True)
        style.setHorizontalAlignment(TextAlignmentType.CENTER)
        style.setVerticalAlignment(TextAlignmentType.CENTER)
        cell.setStyle(style)

    set_cell_value(shenzhn_sheet,5,4,data['shipper_name'])
    set_cell_value(shenzhn_sheet,5,17,data['master_bill_no'])

    set_cell_value(shenzhn_sheet,9,4,data['receiver_name'])
    set_cell_value(shenzhn_sheet,16,4,data['receiver_name'])
    set_cell_value(shenzhn_sheet,26,10,data['total_boxes'])
    set_cell_value(shenzhn_sheet,26,15,data['all_english_name'])
    set_cell_value(shenzhn_sheet,26,22,data['gross_weight'])
    set_cell_value(shenzhn_sheet,26,25,data['volume'])

    set_cell_value(shenzhn_sheet,48,16,data['gross_weight'])
    set_cell_value(shenzhn_sheet,48,19,data['volume'])

    ids = []
    for i in wb.getWorksheets():
        origin_sheetname = i.getName()
        if origin_sheetname == "S#-SZ-customs":
            ids.append(i.getIndex())
    new_SheetSet = SheetSet(ids)

    # 配置PDF保存选项
    saveOption = PdfSaveOptions()
    saveOption.setSheetSet(new_SheetSet)
    # 确保每个工作表单独保存为一个PDF页面
    saveOption.setOnePagePerSheet(True)  # 如果为True，将整个工作表压缩到一个PDF页面上

    # 计算公式
    saveOption.setCalculateFormula(True)  # 计算公式并将其值保存在PDF中

    # 设置字体相关选项
    saveOption.setCheckWorkbookDefaultFont(True)  # 检查工作簿的默认字体，以避免出现方块字符
    saveOption.setCheckFontCompatibility(True)  # 检查每个字符的字体兼容性
    saveOption.setDefaultFont("Arial")  # 设置默认字体（如果未设置正确的字体）

    # 设置图像处理
    saveOption.setImageResample(220, 85)  # 设置图像的PPI和JPEG质量，减少PDF文件大小

    # 设置其他相关选项
    saveOption.setEmbedStandardWindowsFonts(True)  # 嵌入标准的Windows字体
    saveOption.setClearData(False)  # 在保存后不清除工作簿的数据
    saveOption.setCompliance(0)  # 设置PDF标准合规级别，如需要合规的PDF/A等格式
    saveOption.setDisplayDocTitle(True)  # 在PDF窗口的标题栏显示文档标题


    

    
    # 设置PDF文件的完整保存路径
    pdf_file = f"./pdf/customs/{data['master_bill_no']}.pdf"

    # 保存为PDF
    wb.save(pdf_file, saveOption)

# 定义FastAPI接口
@app.post("/process-shipping-data")
async def process_shipping_data_endpoint(request: ShippingRequest, session: Session = Depends(get_session)):
    try:
        results,summary_log = process_shipping_data(
            shipper_name=request.shipper_name,
            receiver_name=request.receiver_name,
            master_bill_no=request.master_bill_no,
            gross_weight=request.gross_weight,
            volume=request.volume,
            product_list=request.product_list,
            totalyugutax = request.totalyugutax,
            predict_tax_price = request.predict_tax_price
        )
        if isinstance(results,dict) and results.get("product_attribute"):
            return JSONResponse({"status":"False","content":"产品属性在地址数据库中不存在"})
        if isinstance(results,dict) and results.get("type")=='net_weight大于gross_weight':
            return JSONResponse({"status":"False","content":results.get("msg")})
        
        shenzhen_cumtoms_data = {
            "shipper_name":request.shipper_name,
            "receiver_name": request.receiver_name,
            "master_bill_no": request.master_bill_no,
            "gross_weight":request.gross_weight,
            "volume":request.volume,
            "total_boxes":summary_log.total_boxes,
            "all_english_name":",".join([i['DESCRIPTION'] for i in results])
        }
        print(shenzhen_cumtoms_data)
        shenzhen_customes_pdf_gennerate(shenzhen_cumtoms_data)
        # excel_path = generate_excel_from_template(results)
        excel_path = generate_excel_from_template_test(results,request.totalyugutax)
        summary_log.packing_type = request.packing_type
        summary_log.port = request.port
        summary_log.filename = Path(excel_path).name
        await create_summary(summary_log,session)

        return FileResponse(path=excel_path, filename=f"{request.master_bill_no} CI&PL.{excel_path.split('.')[-1]}")
    except ValueError as e:
        logger.error(f"Value Error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Value Error: {str(e)}")
    except Exception as e:
        logger.error(f"Internal Server Error: {str(e)}---{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}")

@app.post("/dalei/", response_model=Dalei)
def create_dalei(dalei: DaleiCreate, session: Session = Depends(get_session)):
    new_dalei = Dalei.from_orm(dalei)
    session.add(new_dalei)
    session.commit()
    session.refresh(new_dalei)
    return new_dalei

@app.get("/dalei/")
def read_dalei(skip: int = 0, limit: int = 10, 名称: Optional[str] = None, get_all: bool = False, session: Session = Depends(get_session)):
    query = select(Dalei)
    if 名称:
        query = query.where(Dalei.中文大类.contains(名称))
    total = session.exec(select(func.count()).select_from(query.subquery())).one()
    if get_all:
        dalei_list = session.exec(query).all()
    else:
        dalei_list = session.exec(query.offset(skip).limit(limit)).all()
    return {"items": dalei_list, "total": total}

@app.get("/dalei/{id}", response_model=Dalei)
def read_dalei_by_id(id: int, session: Session = Depends(get_session)):
    dalei = session.get(Dalei, id)
    if not dalei:
        raise HTTPException(status_code=404, detail="Dalei not found")
    return dalei

@app.put("/dalei/{id}", response_model=Dalei)
def update_dalei(id: int, dalei: DaleiCreate, session: Session = Depends(get_session)):
    existing_dalei = session.get(Dalei, id)
    if not existing_dalei:
        raise HTTPException(status_code=404, detail="Dalei not found")
    for key, value in dalei.dict(exclude_unset=True).items():
        setattr(existing_dalei, key, value)
    session.add(existing_dalei)
    session.commit()
    session.refresh(existing_dalei)
    return existing_dalei

@app.delete("/dalei/{id}", response_model=Dalei)
def delete_dalei(id: int, session: Session = Depends(get_session)):
    dalei = session.get(Dalei, id)
    if not dalei:
        raise HTTPException(status_code=404, detail="Dalei not found")
    session.delete(dalei)
    session.commit()
    return dalei
@app.get("/products/", response_model=dict)
def read_products(skip: int = 0, limit: int = 10, 名称: Optional[str] = None,get_all:bool=False,username=None):
    with get_session() as session:
        has_zishui_permission = enforcer.has_permission_for_user(username,'展示自税数据','read','allow')
        if has_zishui_permission:
            query = select(Product3)
        else:
            query = select(Product3).where(Product3.自税==0)
        if 名称:
            query = query.where(Product3.中文品名.contains(名称))
        total = session.exec(select(func.count()).select_from(query.subquery())).one()
        if get_all:
            products = session.exec(query).all()
        else:
            products = session.exec(query.offset(skip).limit(limit)).all()
        return {"items": products, "total": total}

@app.post("/products/", response_model=Product3)
def create_product(product: Product3):
    if not product.更新时间:
        product.更新时间 = datetime.utcnow()
        
    try:
        with get_session() as session:
            session.add(product)
            session.commit()
            try:
                session.refresh(product)
            except InvalidRequestError as e:
                raise HTTPException(status_code=500, detail=f"Error refreshing product: {str(e)}")
            return product
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating product: {str(e)}")

@app.put("/products/{product_id}", response_model=Product3)
def update_product(product_id: int, product: Product3):
    with get_session() as session:
        db_product = session.get(Product3, product_id)
        if not db_product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        # 排除主键字段和None值，并确保字段不为空字符串
        update_data = {
            key: value for key, value in product.dict(exclude_unset=True).items()
            if key != '序号' and value is not None and value != ""
        }
        for key, value in update_data.items():
            setattr(db_product, key, value)

        session.add(db_product)
        session.commit()
        session.refresh(db_product)
        return db_product

@app.delete("/products/{product_id}", response_model=Product3)
def delete_product(product_id: int):
    with get_session() as session:
        db_product = session.get(Product3, product_id)
        if not db_product:
            raise HTTPException(status_code=404, detail="Product not found")
        session.delete(db_product)
        session.commit()
        return db_product

@app.get("/shippersandreceivers/", response_model=dict)
def read_shippers_and_receivers(skip: int = 0, limit: int = 10, ShipperName: Optional[str] = None):
    with get_session() as session:
        query = select(ShippersAndReceivers)
        if ShipperName:
            query = query.where(ShippersAndReceivers.ShipperName.contains(ShipperName))
        total = session.exec(select(func.count()).select_from(query.subquery())).one()
        shippers_and_receivers = session.exec(query.offset(skip).limit(limit)).all()
        return {"items": shippers_and_receivers, "total": total}

@app.post("/shippersandreceivers/", response_model=ShippersAndReceivers)
def create_shipper_or_receiver(shipper_or_receiver: ShippersAndReceivers):
    with get_session() as session:
        session.add(shipper_or_receiver)
        session.commit()
        session.refresh(shipper_or_receiver)
        return shipper_or_receiver

@app.put("/shippersandreceivers/{id}", response_model=ShippersAndReceivers)
def update_shipper_or_receiver(id: int, shipper_or_receiver: ShippersAndReceivers):
    with get_session() as session:
        db_shipper_or_receiver = session.get(ShippersAndReceivers, id)
        if not db_shipper_or_receiver:
            raise HTTPException(status_code=404, detail="Shipper or Receiver not found")
        for key, value in shipper_or_receiver.model_dump(exclude_unset=True).items():
            setattr(db_shipper_or_receiver, key, value)
        session.add(db_shipper_or_receiver)
        session.commit()
        session.refresh(db_shipper_or_receiver)
        return db_shipper_or_receiver

@app.delete("/shippersandreceivers/{id}", response_model=ShippersAndReceivers)
def delete_shipper_or_receiver(id: int):
    with get_session() as session:
        db_shipper_or_receiver = session.get(ShippersAndReceivers, id)
        if not db_shipper_or_receiver:
            raise HTTPException(status_code=404, detail="Shipper or Receiver not found")
        session.delete(db_shipper_or_receiver)
        session.commit()
        return db_shipper_or_receiver

@app.post("/ports/", response_model=Port)
def create_port(port: Port, session: Session = Depends(get_session)):
    session.add(port)
    session.commit()
    session.refresh(port)
    return port

@app.get("/ports/", response_model=List[Port])
def read_ports(session: Session = Depends(get_session), skip: int = 0, limit: Optional[int] = None):
    query = select(Port).offset(skip)
    if limit:
        query = query.limit(limit)
    ports = session.exec(query).all()
    return ports

@app.get("/ports/{port_id}", response_model=Port)
def read_port(port_id: int, session: Session = Depends(get_session)):
    port = session.get(Port, port_id)
    if not port:
        raise HTTPException(status_code=404, detail="Port not found")
    return port

@app.put("/ports/{port_id}", response_model=Port)
def update_port(port_id: int, updated_port: Port, session: Session = Depends(get_session)):
    port = session.get(Port, port_id)
    if not port:
        raise HTTPException(status_code=404, detail="Port not found")
    port.port_name = updated_port.port_name
    port.sender_name = updated_port.sender_name
    port.receiver_name = updated_port.receiver_name
    session.commit()
    session.refresh(port)
    return port

@app.delete("/ports/{port_id}", response_model=Port)
def delete_port(port_id: int, session: Session = Depends(get_session)):
    port = session.get(Port, port_id)
    if not port:
        raise HTTPException(status_code=404, detail="Port not found")
    session.delete(port)
    session.commit()
    return port
# 工厂数据CRUD操作
@app.post("/factory/", response_model=FactoryData)
def create_factory(factory: FactoryData):
    with get_session() as session:
        session.add(factory)
        session.commit()
        session.refresh(factory)
        return factory

@app.get("/factory/")
def read_factories(skip: int = 0, limit: Optional[int] = None):
    with get_session() as session:
        total = session.exec(select(func.count(FactoryData.id))).one()
        query = select(FactoryData).offset(skip)
        if limit is not None:
            query = query.limit(limit)
        factories = session.exec(query).all()
        return {"items": factories, "total": total}

@app.get("/factory/{factory_id}", response_model=FactoryData)
def read_factory(factory_id: int):
    with get_session() as session:
        factory = session.get(FactoryData, factory_id)
        if not factory:
            raise HTTPException(status_code=404, detail="Factory not found")
        return factory

@app.put("/factory/{factory_id}", response_model=FactoryData)
def update_factory(factory_id: int, factory: FactoryData):
    with get_session() as session:
        db_factory = session.get(FactoryData, factory_id)
        if not db_factory:
            raise HTTPException(status_code=404, detail="Factory not found")
        db_factory.属性 = factory.属性
        db_factory.中文名字 = factory.中文名字
        db_factory.英文 = factory.英文
        db_factory.地址 = factory.地址
        session.commit()
        session.refresh(db_factory)
        return db_factory

@app.delete("/factory/{factory_id}", response_model=FactoryData)
def delete_factory(factory_id: int):
    with get_session() as session:
        factory = session.get(FactoryData, factory_id)
        if not factory:
            raise HTTPException(status_code=404, detail="Factory not found")
        session.delete(factory)
        session.commit()
        return factory

# 收发货人CRUD操作
@app.post("/consignee/", response_model=ConsigneeData)
def create_consignee(consignee: ConsigneeData):
    with get_session() as session:
        session.add(consignee)
        session.commit()
        session.refresh(consignee)
        return consignee

@app.get("/consignee/")
def read_consignees(skip: int = 0, limit: Optional[int] = None):
    with get_session() as session:
        total = session.exec(select(func.count(ConsigneeData.id))).one()
        query = select(ConsigneeData).offset(skip)
        if limit is not None:
            query = query.limit(limit)
        consignees = session.exec(query).all()
        return {"items": consignees, "total": total}

@app.get("/consignee/{consignee_id}", response_model=ConsigneeData)
def read_consignee(consignee_id: int):
    with get_session() as session:
        consignee = session.get(ConsigneeData, consignee_id)
        if not consignee:
            raise HTTPException(status_code=404, detail="Consignee not found")
        return consignee

@app.put("/consignee/{consignee_id}", response_model=ConsigneeData)
def update_consignee(consignee_id: int, consignee: ConsigneeData):
    with get_session() as session:
        db_consignee = session.get(ConsigneeData, consignee_id)
        if not db_consignee:
            raise HTTPException(status_code=404, detail="Consignee not found")
        db_consignee.中文 = consignee.中文
        db_consignee.发货人 = consignee.发货人
        db_consignee.发货人详细地址 = consignee.发货人详细地址
        db_consignee.类型 = consignee.类型
        db_consignee.备注 = consignee.备注
        db_consignee.hide = consignee.hide

        session.commit()
        session.refresh(db_consignee)
        return db_consignee

@app.delete("/consignee/{consignee_id}", response_model=ConsigneeData)
def delete_consignee(consignee_id: int):
    with get_session() as session:
        consignee = session.get(ConsigneeData, consignee_id)
        if not consignee:
            raise HTTPException(status_code=404, detail="Consignee not found")
        session.delete(consignee)
        session.commit()
        return consignee
    

@app.get("/api/exchange-rate")
def get_exchange_rate():
    try:
        url = "https://finance.pae.baidu.com/selfselect/sug?wd=%E7%BE%8E%E5%85%83%E4%BA%BA%E6%B0%91%E5%B8%81&skip_login=1&finClientType=pc"


        response = httpx.get(url)
        response.raise_for_status()  # Check for HTTP errors
        data = response.json()
        # 提取汇率
        if data["ResultCode"] == '0':
            cn_us_rate = [
                i['price']
                for i in data['Result']['stock']
                if i['code'] == "USDCNY" or i['code'] == "CNYUSD"
            ]
            
            if not cn_us_rate:
                raise ValueError("Exchange rate not found")
            
            return {"USDCNY":cn_us_rate[0]}

        logger.warning(f"百度获取汇率失败->{data}")
        logger.warning("从外汇网站获取汇率")

        # 获取当天的日期
        today = datetime.today().strftime('%Y-%m-%d')

        # 构建 URL
        url = f"http://m.safe.gov.cn/AppStructured/hlw/jsonRmb.do?date={today}"

        response = httpx.get(url)
        if response.status_code == 200:
            data = response.json()
            exchange_rate = None
            for row in data:
                if row[1] == "美元" and row[3] == '人民币':
                    exchange_rate = round(float(row[2])/row[0],3)
                    return {"USDCNY":exchange_rate}
        
    except requests.RequestException as e:
        raise HTTPException(status_code=503, detail="Service unavailable") from e
    except (KeyError, ValueError) as e:
        raise HTTPException(status_code=500, detail="Error parsing exchange rate data") from e
    
@app.post("/login")
async def login_for_access_token(user: UserLogin, session: Session = Depends(get_session)):
    statement = select(User).where(User.username == user.username)
    result = session.exec(statement)
    user_db = result.first()
    if not user_db or not bcrypt.checkpw(user.password.encode('utf-8'), user_db.password.encode('utf-8')):
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    
    # Assuming the permissions are part of the user_db model
    permissions = enforcer.get_filtered_policy(0, user_db.username)
    
    access_token_expires = timedelta(hours=1)
    access_token = create_access_token(
        data={
            "sub": user_db.username,
            "permissions": permissions
        },
        expires_delta=access_token_expires
    )
    
    refresh_token = create_refresh_token(data={"sub": user_db.username})

    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
    }
@app.post("/users/", response_model=User)
def create_user(user: UserCreate, session: Session = Depends(get_session)):
    statement = select(User).where(User.username == user.username)
    result = session.exec(statement)
    user_db = result.first()
    if user_db:
        raise HTTPException(status_code=400, detail="Username already exists")

    hashed_password = bcrypt.hashpw(user.password.encode('utf-8'), bcrypt.gensalt())
    new_user = User(username=user.username, password=hashed_password.decode('utf-8'))
    session.add(new_user)
    session.commit()
    session.refresh(new_user)

    # 添加用户权限
    for perm in user.permissions:
        obj, act = perm.split(':')
        enforcer.add_policy(user.username, obj, act,'allow')

    enforcer.load_policy()  # 重新加载策略
    return new_user
@app.put("/users/{user_id}/", response_model=User)
def update_user(user_id: int, user_update: UserUpdate, session: Session = Depends(get_session)):
    user_db = session.get(User, user_id)
    if not user_db:
        raise HTTPException(status_code=404, detail="User not found")

    # 更新用户名和密码
    if user_update.username:
        user_db.username = user_update.username
    if user_update.password:
        hashed_password = bcrypt.hashpw(user_update.password.encode('utf-8'), bcrypt.gensalt())
        user_db.password = hashed_password.decode('utf-8')

    # 更新用户权限
    if user_update.permissions is not None:
        current_policies = enforcer.get_filtered_policy(0, user_db.username)
        current_permissions = {f"{policy[1]}:{policy[2]}" for policy in current_policies if policy[3] == 'allow'}
        update_permissions = set(user_update.permissions)
        print(f"当前权限: {current_permissions}")
        print(f"更新权限: {update_permissions}")
        
        # 需要删除的权限（将 eft 设置为 deny）
        for perm in current_permissions - update_permissions:
            print(f"删除权限{perm}")
            obj, act = perm.split(':')
            enforcer.update_policy([user_db.username, obj, act, "allow"], [user_db.username, obj, act, "deny"])

        # 需要添加的权限
        for perm in update_permissions - current_permissions:
            obj, act = perm.split(':')
            enforcer.add_policy(user_db.username, obj, act, "allow")

    session.add(user_db)
    session.commit()
    session.refresh(user_db)

    enforcer.load_policy()  # 重新加载策略
    return user_db
@app.get("/users/", response_model=List[User])
def read_users(skip: int = 0, limit: int = 10, session: Session = Depends(get_session)):
    users = session.exec(select(User).offset(skip).limit(limit)).all()
    return users

@app.get("/users/{user_id}/", response_model=User)
def read_user(user_id: int, session: Session = Depends(get_session)):
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user
@app.delete("/users/{user_id}/", response_model=User)
def delete_user(user_id: int, session: Session = Depends(get_session)):
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # 删除用户权限
    enforcer.delete_roles_for_user(user.username)
    enforcer.delete_user(user.username)
    
    session.delete(user)
    session.commit()
    
    return {"message": "User and associated permissions deleted successfully"}
@app.post("/add_policy/")
async def add_policy(policy: Policy):
    if enforcer.add_policy(policy.sub, policy.obj, policy.act, policy.eft):
        enforcer.load_policy()  # 重新加载策略
        return {"message": "策略添加成功"}
    else:
        raise HTTPException(status_code=400, detail="策略已存在或无法添加")

@app.delete("/remove_policy/")
async def remove_policy(policy: Policy):
    if enforcer.remove_policy(policy.sub, policy.obj, policy.act, policy.eft,"",""):
        enforcer.load_policy()  # 重新加载策略
        return {"message": "策略删除成功"}
    else:
        raise HTTPException(status_code=400, detail="策略不存在或无法删除")

@app.put("/update_policy/")
async def update_policy(update_policy: UpdatePolicy):
    old_policy = [update_policy.old_sub, update_policy.old_obj, update_policy.old_act, update_policy.old_eft,"",""]
    new_policy = [update_policy.new_sub, update_policy.new_obj, update_policy.new_act, update_policy.new_eft,"",""]

    # 检查旧策略是否存在
    if not enforcer.has_policy(*old_policy):
        raise HTTPException(status_code=404, detail="旧策略不存在")
    


    # 尝试更新策略
    result = enforcer.update_policy(old_policy, new_policy)

    if result:
        enforcer.load_policy()  # 重新加载策略
        return {"message": "策略更新成功"}
    else:
        raise HTTPException(status_code=400, detail="策略更新失败或未找到旧策略")


@app.get("/get_policies/")
async def get_policies():
    return enforcer.get_policy()

@app.get("/get_user_policies/")
async def get_user_policies(user: str):
    user_policies = enforcer.get_filtered_policy(0, user)
    if not user_policies:
        raise HTTPException(status_code=404, detail="该用户没有策略")

    return user_policies

@app.post("/add_group/")
async def add_group(group: Group):
    if enforcer.add_grouping_policy(group.user, group.group):
        enforcer.load_policy()  # 重新加载策略
        return {"message": "组添加成功"}
    else:
        raise HTTPException(status_code=400, detail="组已存在或无法添加")

@app.delete("/remove_group/")
async def remove_group(group: Group):
    if enforcer.remove_grouping_policy(group.user, group.group):
        enforcer.load_policy()  # 重新加载策略
        return {"message": "组删除成功"}
    else:
        raise HTTPException(status_code=400, detail="组不存在或无法删除")

@app.get("/get_groups/")
async def get_groups():
    return enforcer.get_grouping_policy()


@app.get("/get-ip/")
async def get_ip(request: Request):
    client_host = request.client.host
    return {"ip": client_host}


# 创建IP白名单
@app.post("/ip_white_list/", response_model=IpWhiteList)
def create_ip_white_list(ip_white_list: IpWhiteList, session: Session = Depends(get_session)):
    try:
        db_ip_white_list = session.exec(select(IpWhiteList).where(IpWhiteList.ip == ip_white_list.ip)).first()
        if db_ip_white_list:
            raise HTTPException(status_code=400, detail="IP already exists")
        session.add(ip_white_list)
        session.commit()
        session.refresh(ip_white_list)
        return ip_white_list
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# 获取所有IP白名单
@app.get("/ip_white_list/", response_model=List[IpWhiteList])
def get_all_ip_white_list(session: Session = Depends(get_session)):
    ip_white_lists = session.exec(select(IpWhiteList)).all()
    return ip_white_lists

# 获取单个IP白名单
@app.get("/ip_white_list/{ip_white_list_id}", response_model=IpWhiteList)
def get_ip_white_list(ip_white_list_id: int, session: Session = Depends(get_session)):
    ip_white_list = session.get(IpWhiteList, ip_white_list_id)
    if not ip_white_list:
        raise HTTPException(status_code=404, detail="IP white list not found")
    return ip_white_list

# 更新IP白名单
@app.put("/ip_white_list/{ip_white_list_id}", response_model=IpWhiteList)
def update_ip_white_list(ip_white_list_id: int, ip_white_list: IpWhiteList, session: Session = Depends(get_session)):
    db_ip_white_list = session.get(IpWhiteList, ip_white_list_id)
    if not db_ip_white_list:
        raise HTTPException(status_code=404, detail="IP white list not found")
    ip_white_list_data = ip_white_list.dict(exclude_unset=True)
    for key, value in ip_white_list_data.items():
        setattr(db_ip_white_list, key, value)
    session.add(db_ip_white_list)
    session.commit()
    session.refresh(db_ip_white_list)
    return db_ip_white_list

# 删除IP白名单
@app.delete("/ip_white_list/{ip_white_list_id}", response_model=IpWhiteList)
def delete_ip_white_list(ip_white_list_id: int, session: Session = Depends(get_session)):
    ip_white_list = session.get(IpWhiteList, ip_white_list_id)
    if not ip_white_list:
        raise HTTPException(status_code=404, detail="IP white list not found")
    session.delete(ip_white_list)
    session.commit()
    return ip_white_list


@app.get("/files", response_model=List[FileInfo])
async def get_files():
    directory_path = "./pdf"
    if not os.path.exists(directory_path):
        raise HTTPException(status_code=404, detail="Directory not found")
    
    files = []
    for file_name in os.listdir(directory_path):
        file_path = os.path.join(directory_path, file_name)
        if os.path.isfile(file_path):
            file_info = os.stat(file_path)
            files.append(FileInfo(name=file_name, time=datetime.fromtimestamp(file_info.st_mtime)))
    
    files.sort(key=lambda x: x.time, reverse=True)
    
    return files[:500]

@app.get("/download/{file_name}")
async def download_file(file_name: str):
    file_path = os.path.join("./pdf", file_name)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(file_path, media_type='application/octet-stream', filename=file_name)



@app.post("/haiyunzishui/", response_model=HaiYunZiShui)
def create_haiyunzishui(haiyunzishui: HaiYunZiShui, session: Session = Depends(get_session)):
    session.add(haiyunzishui)
    session.commit()
    session.refresh(haiyunzishui)
    return haiyunzishui

@app.get("/haiyunzishui/", response_model=List[HaiYunZiShui])
def read_haiyunzishuis(session: Session = Depends(get_session), skip: int = 0, limit: Optional[int] = None):
    query = select(HaiYunZiShui).offset(skip)
    if limit:
        query = query.limit(limit)
    haiyunzishui = session.exec(query).all()
    return haiyunzishui

@app.get("/haiyunzishui/{haiyunzishui_id}", response_model=HaiYunZiShui)
def read_haiyunzishui(haiyunzishui_id: int, session: Session = Depends(get_session)):
    haiyunzishui = session.get(HaiYunZiShui, haiyunzishui_id)
    if not haiyunzishui:
        raise HTTPException(status_code=404, detail="Port not found")
    return haiyunzishui

@app.put("/haiyunzishui/{haiyunzishui_id}", response_model=HaiYunZiShui)
def update_haiyunzishui(haiyunzishui_id: int, updated_port: HaiYunZiShui, session: Session = Depends(get_session)):
    haiyunzishui = session.get(HaiYunZiShui, haiyunzishui_id)
    if not haiyunzishui:
        raise HTTPException(status_code=404, detail="Port not found")
    haiyunzishui.zishui_name = updated_port.zishui_name
    haiyunzishui.sender = updated_port.sender
    haiyunzishui.receiver = updated_port.receiver
    session.commit()
    session.refresh(haiyunzishui)
    return haiyunzishui

@app.delete("/haiyunzishui/{haiyunzishui_id}", response_model=HaiYunZiShui)
def delete_haiyunzishui(haiyunzishui_id: int, session: Session = Depends(get_session)):  # noqa: F811
    haiyunzishui = session.get(HaiYunZiShui, haiyunzishui_id)
    if not haiyunzishui:
        raise HTTPException(status_code=404, detail="Port not found")
    session.delete(haiyunzishui)
    session.commit()
    return haiyunzishui


class CustomClearHistorySummaryLog(SQLModel, table=True):
    __tablename__ = "custom_clear_history_summary_log"
    id: UUID = Field(default_factory=uuid4, primary_key=True)
    filename: str
    generation_time: datetime = Field(default_factory=datetime.utcnow)
    port: str
    packing_type: str
    shipper: str
    consignee: str
    estimated_tax_amount: float
    gross_weight_kg: float
    volume_cbm: float
    total_boxes: int
    estimated_tax_rate_cny_per_kg: float
    remarks:str

    details: List["CustomClearHistoryDetailLog"] = Relationship(back_populates="summary")

class CustomClearHistoryDetailLog(SQLModel, table=True):
    __tablename__ = "custom_clear_history_detail_log"
    id: Optional[int] = Field(default=None, primary_key=True)
    hs_code: str
    chinese_name: str
    transport_mode: str
    master_bill_number: str
    generation_time: datetime = Field(default_factory=datetime.utcnow)
    total_tax_rate: float
    exemption_code: str
    category: str

    summary_log_id: UUID = Field(sa_column=Column(SQLAlchemyUUID(as_uuid=True), ForeignKey("custom_clear_history_summary_log.id")))
    summary: CustomClearHistorySummaryLog = Relationship(back_populates="details")


# 创建 Summary 记录
@app.post("/cumstom_clear_history_summary/", response_model=CustomClearHistorySummaryLog)
async def create_summary(summary: CustomClearHistorySummaryLog, session: Session = Depends(get_session)):
    try:
        # 开始事务
        with session.begin():
            # 添加主记录
            session.add(summary)
            session.flush()  # 刷新会话以获取生成的 ID

            # 更新详细记录
            for detail in summary.details:
                detail.summary_log_id = summary.id
                detail.generation_time = summary.generation_time
                session.add(detail)

        # 提交事务
        session.commit()
        session.refresh(summary)
        return summary
    except Exception as e:
        # 回滚事务
        session.rollback()
        raise e
@app.post("/update_cumstom_clear_history_summary_remarks/", response_model=CustomClearHistorySummaryLog)
async def update_summary(request_body: update_cumstom_clear_history_summary_remarks, session: Session = Depends(get_session)):
    summary = session.get(CustomClearHistorySummaryLog, request_body.id)
    summary.remarks = request_body.remarks
    session.add(summary)
    session.commit()
    session.refresh(summary)

    return summary
class SummaryResponse(BaseModel):
    summaries: List[CustomClearHistorySummaryLog]
    total: Optional[int] = None
    total_pages: Optional[int] = None

@app.get("/cumstom_clear_history_summary/", response_model=SummaryResponse)
def read_summaries(
    enable_pagination: bool = Query(False, description="Enable pagination"),
    page: int = Query(1, description="Page number", ge=1),
    page_size: int = Query(10, description="Number of items per page", ge=1, le=100),
    file_name: Optional[str] = Query(None, description="File name to filter by"),
    session: Session = Depends(get_session)
):
    query = select(CustomClearHistorySummaryLog).order_by(CustomClearHistorySummaryLog.generation_time.desc())
    
    # 如果 file_name 有值，添加 WHERE 约束
    if file_name:
        query = query.filter(CustomClearHistorySummaryLog.filename.like(f"%{file_name}%"))
    
    if enable_pagination:
        # 计算偏移量
        offset = (page - 1) * page_size
        
        # 查询分页数据
        summaries = session.exec(
            query.offset(offset).limit(page_size)
        ).all()
        
        # 查询总记录数
        total_query = select(func.count()).select_from(query.subquery())
        total_result = session.exec(total_query).first()
        total = total_result
        
        # 计算总页数
        total_pages = (total + page_size - 1) // page_size
        
        return {
            "summaries": summaries,
            "total": total,
            "total_pages": total_pages
        }
    else:
        # 查询全部数据
        summaries = session.exec(query).all()
        
        return {
            "summaries": summaries,
            "total":summaries.count,
            "total_pages": 1
        }



# 查询单个 Summary 记录
@app.get("/cumstom_clear_history_summary/{summary_id}", response_model=CustomClearHistorySummaryLog)
def read_summary(summary_id: UUID, session: Session = Depends(get_session)):
    summary = session.get(CustomClearHistorySummaryLog, summary_id)
    if not summary:
        raise HTTPException(status_code=404, detail="Summary not found")
    return summary

# 创建 Detail 记录
@app.post("/cumstom_clear_history_detail/", response_model=CustomClearHistoryDetailLog)
def create_detail(detail: CustomClearHistoryDetailLog, session: Session = Depends(get_session)):
    session.add(detail)
    session.commit()
    session.refresh(detail)
    return detail

# 查询所有 Detail 记录
@app.get("/cumstom_clear_history_detail/", response_model=List[CustomClearHistoryDetailLog])
def read_details(session: Session = Depends(get_session)):
    details = session.exec(select(CustomClearHistoryDetailLog)).all()
    return details

# 查询单个 Detail 记录
@app.get("/cumstom_clear_history_detail/{detail_id}", response_model=CustomClearHistoryDetailLog)
def read_detail(detail_id: int, session: Session = Depends(get_session)):
    detail = session.get(CustomClearHistoryDetailLog, detail_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Detail not found")
    return detail

@app.get("/output_cumtoms_clear_log/")
def output_log(
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
     session: Session = Depends(get_session)
):
    # 查询汇总日志
    summary_logs = session.exec(
        select(CustomClearHistorySummaryLog).where(
            CustomClearHistorySummaryLog.generation_time >= start_time,
            CustomClearHistorySummaryLog.generation_time <= end_time
        )
    ).all()

    # 查询明细日志
    detail_logs = session.exec(
        select(CustomClearHistoryDetailLog).where(
            CustomClearHistoryDetailLog.generation_time >= start_time,
            CustomClearHistoryDetailLog.generation_time <= end_time
        )
    ).all()

    # 创建 Excel 文件
    wb = Openpyxl_Workbook()
    ws_summary = wb.active
    ws_summary.title = "历史汇总"
    ws_detail = wb.create_sheet(title="历史明细")

    # 写入汇总日志数据
    summary_headers = ["ID", "文件名", "生成时间", "港口", "装箱类型", "发货人", "收货人", "预估税金", "Gross Weight (kg)", "Volume (cbm)", "总箱数", "预估税金单价CNY/Kg", "Remarks"]
    ws_summary.append(summary_headers)
    for log in summary_logs:
        ws_summary.append([
            str(log.id), log.filename, log.generation_time.isoformat(), log.port, log.packing_type, log.shipper, log.consignee, log.estimated_tax_amount, log.gross_weight_kg, log.volume_cbm, log.total_boxes, log.estimated_tax_rate_cny_per_kg, log.remarks
        ])

    # 写入明细日志数据
    detail_headers = ["ID", "HS Code", "中文品名", "运输方式", "主单号", "生成时间", "总税率", "豁免代码", "类别","FZ", "Summary Log ID"]
    ws_detail.append(detail_headers)
    for log in detail_logs:
        ws_detail.append([
            log.id, log.hs_code, log.chinese_name, log.transport_mode, log.master_bill_number, log.generation_time.isoformat(), log.total_tax_rate, log.exemption_code, log.category, 1,str(log.summary_log_id)
        ])

    # 将 Excel 文件保存到内存中
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回 Excel 文件
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=custom_clear_history_log.xlsx"})

class ShipmentLog(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    shipper_name: str
    receiver_name: str
    master_bill_no: str
    gross_weight: float
    volume: float
    total_boxes: int
    all_english_name: str = Field(sa_column=Column(Text))  # 使用 SQLAlchemy 的 Text 类型
    status: int = Field(default=0)  # 0: 未完成, 1: 已完成, -1: 失败

# 增：创建 ShipmentLog
@app.post("/shipment_logs/", response_model=ShipmentLog)
async def create_shipment_log(shipment_log: ShipmentLog, session: Session = Depends(get_session)):
    session.add(shipment_log)
    await session.commit()
    await session.refresh(shipment_log)
    return shipment_log

# 改：更新 ShipmentLog
@app.put("/shipment_logs/{shipment_log_id}", response_model=ShipmentLog)
async def update_shipment_log(shipment_log_id: int, shipment_log: ShipmentLog, session: Session = Depends(get_session)):
    db_shipment_log = await session.get(ShipmentLog, shipment_log_id)
    if not db_shipment_log:
        raise HTTPException(status_code=404, detail="ShipmentLog not found")
    shipment_log_data = shipment_log.dict(exclude_unset=True)
    for key, value in shipment_log_data.items():
        setattr(db_shipment_log, key, value)
    session.add(db_shipment_log)
    await session.commit()
    await session.refresh(db_shipment_log)
    return db_shipment_log

# 查：查询所有 ShipmentLog
@app.get("/shipment_logs/", response_model=list[ShipmentLog])
async def read_shipment_logs(
    status: Optional[int] = Query(None, description="Filter by status"),
    offset: int = Query(0, description="Offset for pagination"),
    limit: int = Query(10, description="Limit for pagination"),
    session: Session = Depends(get_session)
):
    query = select(ShipmentLog)
    if status is not None:
        query = query.where(ShipmentLog.status == status)
    query = query.offset(offset).limit(limit)
    result = await session.execute(query)
    shipment_logs = result.scalars().all()
    return shipment_logs

# 查：根据 ID 查询 ShipmentLog
@app.get("/shipment_logs/{shipment_log_id}", response_model=ShipmentLog)
async def read_shipment_log(shipment_log_id: int, session: Session = Depends(get_session)):
    shipment_log = await session.get(ShipmentLog, shipment_log_id)
    if not shipment_log:
        raise HTTPException(status_code=404, detail="ShipmentLog not found")
    return shipment_log
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app=app,host="0.0.0.0",port=8085)

